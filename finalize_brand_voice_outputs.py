# -*- coding: utf-8 -*-
"""Rewrite cleaned brand-voice outputs into customer-facing marketing copy."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import Alignment


INPUT_FILE = "01_train_examples_brand_voice_UTF8_CLEANED.xlsx"
OUTPUT_FILE = "01_train_examples_brand_voice_UTF8_FINAL.xlsx"
REQUIRED_COLUMNS = ["input", "output", "industry", "tone", "type", "product"]

META_PHRASES = [
    "Nội dung tập trung vào",
    "Nội dung này",
    "Chỉ nhắc đến",
    "không thêm tính năng mới",
    "Sản phẩm được mô tả",
    "Thông điệp hướng về",
    "Điểm nổi bật đầu tiên",
    "Điểm tiếp theo cần nhắc",
    "#BaiDangMangXa",
    "từ prompt",
    "prompt",
    "thông tin đầu vào",
    "dữ liệu đã có",
    "tiêu chí",
    "format",
    "field",
    "cấu trúc bài",
    "bài viết này",
    "email này",
    "được trình bày như",
    "được giới thiệu theo",
    "giúp có",
]

FORBIDDEN_GENERAL = [
    "Góc cần biết",
    "Dưới đây là",
    "Tôi đã sửa",
    "cam kết 100%",
    "tốt nhất thị trường",
    "chữa khỏi",
    "đảm bảo lợi nhuận",
    "duy nhất hôm nay",
    "sắp hết hàng",
]

URGENT_WEAK = ["không cần mua vội", "lưu lại để xem sau", "cân nhắc từ từ"]

TYPE_FIELDS = {
    "headline": ["Headline", "Subheadline", "Lời kêu gọi hành động"],
    "description": ["Mô tả ngắn", "Lợi ích chính", "Đặc điểm nổi bật", "Lời kêu gọi hành động"],
    "social": ["Hook", "Caption", "Lời kêu gọi hành động", "Hashtags"],
    "email": ["Subject", "Preview text", "Lời chào", "Nội dung chính", "Lời kêu gọi hành động"],
    "cta": ["Lời kêu gọi hành động chính", "Microcopy"],
}

TYPE_VERSION_COUNTS = {
    "headline": 3,
    "description": 3,
    "social": 3,
    "email": 3,
    "cta": 3,
}

PRODUCT_PROFILES = {
    "bình giữ nhiệt thép 316": {
        "scene": "mang cà phê từ nhà đến văn phòng",
        "activity": "mang cà phê đi làm",
        "problem": "cà phê nhanh nguội hoặc bình rò nước trong túi",
        "result": "đồ uống vẫn sẵn sàng khi bạn bước vào buổi họp tiếp theo",
        "comfort": "yên tâm mở túi mà không lo giấy tờ bị ướt",
        "daily": "mỗi sáng đi làm",
        "hashtags": ["#BinhGiuNhiet", "#CaPheDiLam", "#DanVanPhong", "#DoDungTienIch"],
    },
    "đèn bàn LED chống chói": {
        "scene": "học khuya, đọc tài liệu hoặc làm việc tại nhà",
        "activity": "học khuya và làm việc tại nhà",
        "problem": "ánh sáng không hợp làm mắt nhanh mỏi",
        "result": "góc bàn sáng vừa đủ để tập trung lâu hơn",
        "comfort": "dễ chỉnh ánh sáng theo từng việc đang làm",
        "daily": "mỗi buổi tối học tập hoặc làm việc",
        "hashtags": ["#DenBanLED", "#GocHocTap", "#LamViecTaiNha", "#HocKhuya"],
    },
    "bộ chăm sóc da tối giản": {
        "scene": "bắt đầu skincare mà không muốn quá nhiều bước",
        "activity": "chăm da mỗi ngày",
        "problem": "routine rườm rà khiến người mới dễ bỏ cuộc",
        "result": "việc chăm da hằng ngày nhẹ nhàng và dễ duy trì hơn",
        "comfort": "cảm giác da được chăm chút mà không mất quá nhiều thời gian",
        "daily": "mỗi sáng và tối",
        "hashtags": ["#SkincareToiGian", "#ChamSocDa", "#RoutineDonGian", "#NguoiMoiSkincare"],
    },
    "máy xay sinh tố mini": {
        "scene": "chuẩn bị sinh tố nhanh trước khi ra ngoài",
        "activity": "chuẩn bị đồ uống nhanh",
        "problem": "buổi sáng bận rộn, không muốn bày nhiều đồ rồi dọn rửa lâu",
        "result": "bạn có ly đồ uống cá nhân gọn gàng để mang theo",
        "comfort": "xay xong cầm đi ngay, bếp vẫn gọn",
        "daily": "những buổi sáng ít thời gian",
        "hashtags": ["#MayXayMini", "#SinhToMangDi", "#SongBanRon", "#DoUongNhanh"],
    },
    "ghế công thái học lưng lưới": {
        "scene": "ngồi máy tính nhiều giờ ở văn phòng hoặc tại nhà",
        "activity": "ngồi làm việc lâu",
        "problem": "lưng mỏi và khó giữ tư thế khi làm việc lâu",
        "result": "chỗ ngồi thoáng hơn, đỡ lưng tốt hơn trong ngày dài",
        "comfort": "ngồi lâu vẫn dễ chịu hơn nhờ lưng lưới và tay ghế chỉnh được",
        "daily": "trong ca làm việc dài",
        "hashtags": ["#GheCongThaiHoc", "#GocLamViec", "#LamViecVanPhong", "#NgoiThoaiMai"],
    },
    "nồi chiên không dầu dung tích 6 lít": {
        "scene": "chuẩn bị bữa ăn gia đình nhanh gọn",
        "activity": "nấu bữa tối nhanh gọn",
        "problem": "nấu món nóng mất thời gian và ngại lau rửa sau bữa ăn",
        "result": "bữa cơm có món nóng giòn mà bếp vẫn gọn hơn",
        "comfort": "hẹn giờ rồi làm việc khác trong lúc chờ món chín",
        "daily": "những bữa tối bận rộn",
        "hashtags": ["#NoiChienKhongDau", "#BepGiaDinh", "#NauAnNhanh", "#MonNgonMoiNgay"],
    },
    "balo laptop chống nước": {
        "scene": "đi làm, gặp khách hàng hoặc di chuyển nhiều điểm trong ngày",
        "activity": "mang laptop đi làm",
        "problem": "mưa bất chợt và balo nặng làm việc mang laptop kém thoải mái",
        "result": "laptop và đồ cá nhân được sắp gọn hơn khi ra ngoài",
        "comfort": "vai đỡ mỏi hơn khi phải mang laptop cả ngày",
        "daily": "mỗi lần ra khỏi nhà đi làm",
        "hashtags": ["#BaloLaptop", "#DiLamMoiNgay", "#ChongNuoc", "#PhuKienCongSo"],
    },
    "tai nghe không dây chống ồn": {
        "scene": "làm việc tập trung, họp online hoặc di chuyển trong không gian ồn",
        "activity": "làm việc tập trung",
        "problem": "tiếng ồn xung quanh làm bạn khó tập trung và nghe gọi kém rõ",
        "result": "giữ được nhịp làm việc riêng ngay cả khi xung quanh nhiều âm thanh",
        "comfort": "đeo lên là có khoảng yên tĩnh cho cuộc gọi hoặc bản nhạc quen",
        "daily": "trong giờ làm việc và khi di chuyển",
        "hashtags": ["#TaiNgheChongOn", "#LamViecTapTrung", "#HopOnline", "#Pin30Gio"],
    },
    "khóa cửa thông minh": {
        "scene": "quản lý việc ra vào căn hộ cho gia đình hoặc khách ngắn hạn",
        "activity": "ra vào căn hộ",
        "problem": "quên chìa khóa hoặc khó chia sẻ quyền vào nhà đúng lúc",
        "result": "ra vào chủ động hơn mà vẫn dễ kiểm soát",
        "comfort": "mở cửa nhanh bằng vân tay và an tâm hơn khi cấp mã tạm thời",
        "daily": "mỗi lần ra vào căn hộ",
        "hashtags": ["#KhoaCuaThongMinh", "#CanHoHienDai", "#VanTay", "#NhaAnTam"],
    },
    "thảm yoga chống trượt": {
        "scene": "tập yoga, giãn cơ hoặc tập nhẹ tại nhà",
        "activity": "tập tại nhà",
        "problem": "thảm dễ xê dịch làm bạn ngại đổi động tác",
        "result": "buổi tập ổn định hơn từ động tác đầu tiên",
        "comfort": "bề mặt bám tốt và độ dày vừa đủ để gối, cổ tay dễ chịu hơn",
        "daily": "mỗi buổi tập tại nhà",
        "hashtags": ["#ThamYoga", "#TapTaiNha", "#ChongTruot", "#YogaMoiNgay"],
    },
    "máy lọc không khí phòng ngủ": {
        "scene": "giữ phòng ngủ sạch thoáng hơn cho gia đình có trẻ nhỏ",
        "activity": "giữ phòng ngủ sạch thoáng",
        "problem": "bụi trong phòng và tiếng máy ồn làm giấc ngủ kém dễ chịu",
        "result": "không gian nghỉ ngơi yên tĩnh hơn vào ban đêm",
        "comfort": "theo dõi bụi dễ hơn và bật chế độ ngủ khi cả nhà nghỉ ngơi",
        "daily": "mỗi tối trước giờ ngủ",
        "hashtags": ["#MayLocKhongKhi", "#PhongNguSach", "#GiaDinhCoTre", "#GiacNguYenTinh"],
    },
    "ví da mini nhiều ngăn": {
        "scene": "mang thẻ, tiền mặt và giấy tờ nhỏ khi đi làm hoặc đi chơi",
        "activity": "mang thẻ và tiền gọn gàng",
        "problem": "ví cồng kềnh làm túi nặng và thẻ dễ lẫn lộn",
        "result": "đồ cần dùng nằm gọn trong tay, lấy ra nhanh hơn",
        "comfort": "da mềm, nhiều ngăn thẻ và kích thước nhỏ giúp túi xách gọn hơn",
        "daily": "mỗi lần ra ngoài",
        "hashtags": ["#ViDaMini", "#PhuKienGonGang", "#NhieuNganThe", "#KhacTen"],
    },
}


@dataclass
class RowData:
    row_number: int
    input: str
    output: str
    industry: str
    tone: str
    content_type: str
    product: str
    keywords: list[str]
    target: str
    extra: str


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    return unicodedata.normalize("NFC", text).replace("\r\n", "\n").replace("\r", "\n").strip()


def norm_compare(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", value).lower()).strip()


def cap_first(text: str) -> str:
    text = text.strip()
    return text[:1].upper() + text[1:] if text else text


def strip_dot(text: str) -> str:
    return re.sub(r"[\s.]+$", "", text.strip())


def sentence(text: str) -> str:
    text = re.sub(r"\s+", " ", text.strip()).strip(" .;,")
    if text and text[-1] not in ".!?":
        text += "."
    return text


def join_items(items: list[str]) -> str:
    clean = [strip_dot(item) for item in items if strip_dot(item)]
    if not clean:
        return "những chi tiết tiện dụng"
    if len(clean) == 1:
        return clean[0]
    if len(clean) == 2:
        return f"{clean[0]} và {clean[1]}"
    return f"{', '.join(clean[:-1])} và {clean[-1]}"


def find_input_path() -> Path:
    root = Path.cwd()
    preferred = [root / INPUT_FILE, root / "fine_tuning_excel_pack_brand_voice_UTF8_FIXED" / INPUT_FILE]
    for path in preferred:
        if path.exists():
            return path
    matches = sorted(path for path in root.rglob(INPUT_FILE) if ".git" not in path.parts)
    if not matches:
        raise FileNotFoundError(f"Không tìm thấy {INPUT_FILE}")
    return matches[0]


def find_target_sheet(workbook):
    for ws in workbook.worksheets:
        header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_values:
            continue
        header_map = {normalize_text(value).lower(): idx + 1 for idx, value in enumerate(header_values) if value is not None}
        if all(column in header_map for column in REQUIRED_COLUMNS):
            return ws, header_map
    raise ValueError("Không tìm thấy sheet có đủ 6 cột bắt buộc")


def extract_label(input_text: str, label: str) -> str:
    wanted = label.lower() + ":"
    for line in normalize_text(input_text).split("\n"):
        line = line.strip()
        if line.lower().startswith(wanted):
            return strip_dot(line.split(":", 1)[1])
    return ""


def split_keywords(raw: str) -> list[str]:
    keywords = [strip_dot(part) for part in re.split(r"[;,\n]+", raw) if strip_dot(part)]
    return keywords[:3]


def get_profile(product: str) -> dict[str, object]:
    return PRODUCT_PROFILES.get(
        product,
        {
            "scene": "sử dụng hằng ngày",
            "activity": "sử dụng hằng ngày",
            "problem": "việc chọn sản phẩm phù hợp còn mất thời gian",
            "result": "trải nghiệm mua sắm gọn và dễ chịu hơn",
            "comfort": "cảm giác yên tâm hơn khi chọn đúng nhu cầu",
            "daily": "mỗi ngày",
            "hashtags": ["#MuaSamThongMinh", "#DoDungTienIch", "#SongGonGang"],
        },
    )


def read_rows(ws, header_map: dict[str, int]) -> list[RowData]:
    rows: list[RowData] = []
    for row_number in range(2, ws.max_row + 1):
        input_text = normalize_text(ws.cell(row=row_number, column=header_map["input"]).value)
        product = normalize_text(ws.cell(row=row_number, column=header_map["product"]).value)
        keywords = split_keywords(extract_label(input_text, "Từ khóa chính"))
        while len(keywords) < 3:
            keywords.append("dễ dùng hằng ngày")
        rows.append(
            RowData(
                row_number=row_number,
                input=input_text,
                output=normalize_text(ws.cell(row=row_number, column=header_map["output"]).value),
                industry=normalize_text(ws.cell(row=row_number, column=header_map["industry"]).value),
                tone=normalize_text(ws.cell(row=row_number, column=header_map["tone"]).value).lower(),
                content_type=normalize_text(ws.cell(row=row_number, column=header_map["type"]).value).lower(),
                product=product,
                keywords=keywords,
                target=extract_label(input_text, "Đối tượng mục tiêu") or "người đang cần mua sắm thông minh",
                extra=extract_label(input_text, "Thông tin bổ sung"),
            )
        )
    return rows


def extra_phrase(row: RowData) -> str:
    if not row.extra:
        return ""
    return sentence(f"Hiện có {row.extra}")


def has_offer(row: RowData) -> bool:
    return bool(row.extra) and any(token in row.extra.lower() for token in ["freeship", "giảm", "tặng", "ưu đãi", "miễn phí", "đổi", "khắc", "tư vấn", "bảo hành", "kèm"])


def cta(row: RowData, index: int) -> str:
    offer = has_offer(row)
    options = {
        "urgent": [
            "Xem ưu đãi ngay" if offer else "Xem sản phẩm ngay",
            "Đặt mua hôm nay",
            "Nhắn shop để giữ ưu đãi" if offer else "Nhắn shop để được tư vấn ngay",
        ],
        "professional": ["Xem chi tiết sản phẩm", "Yêu cầu tư vấn", "Chọn phương án phù hợp"],
        "friendly": ["Xem thử sản phẩm", "Nhắn shop để được gợi ý", "Chọn mẫu hợp với bạn"],
        "luxury": ["Khám phá lựa chọn tinh tế", "Nhận tư vấn riêng", "Chọn trải nghiệm xứng đáng"],
        "humorous": ["Xem ngay cho đỡ tò mò", "Nhắn shop hỏi nhanh", "Chọn món hợp gu"],
        "emotional": ["Chọn cảm giác yên tâm hơn", "Nhắn shop để được gợi ý", "Xem lựa chọn dành cho bạn"],
    }
    return options.get(row.tone, options["friendly"])[index % 3]


def versioned(parts: list[str]) -> str:
    return "\n\n".join(f"Phiên bản {idx}:\n{part.strip()}" for idx, part in enumerate(parts, start=1))


def build_headline(row: RowData) -> str:
    p = row.product
    pc = cap_first(p)
    profile = get_profile(p)
    scene = profile["scene"]
    activity = profile["activity"]
    problem = profile["problem"]
    result = profile["result"]
    comfort = profile["comfort"]
    f1, f2, f3 = row.keywords[:3]
    bonus = extra_phrase(row)

    if row.tone == "urgent":
        bodies = [
            f"Headline: Cần {p} cho việc {activity}? Xem ngay hôm nay.\nSubheadline: {pc} giúp xử lý chuyện {problem}; có {join_items(row.keywords)} để bạn dùng thuận tiện hơn. {bonus}\nLời kêu gọi hành động: {cta(row, 0)}.",
            f"Headline: Đừng để {problem} làm ngày của bạn kém gọn.\nSubheadline: Chọn {p} với {f1}, {f2} và {f3} để sẵn sàng hơn trong {profile['daily']}.\nLời kêu gọi hành động: {cta(row, 1)}.",
            f"Headline: {pc} đang đáng xem nếu bạn cần mua sớm.\nSubheadline: Một lựa chọn thực tế cho {row.target}, giúp {result}.\nLời kêu gọi hành động: {cta(row, 2)}.",
        ]
    elif row.tone == "friendly":
        bodies = [
            f"Headline: {pc} cho {profile['daily']} nhẹ nhàng hơn.\nSubheadline: Dễ dùng, dễ hỏi thêm và phù hợp với {row.target} đang muốn {result}.\nLời kêu gọi hành động: {cta(row, 0)}.",
            f"Headline: Chọn {p} vừa đủ tiện, không cần phức tạp.\nSubheadline: Bạn có {f1}, thêm {f2} và {f3} để việc {activity} dễ chịu hơn.\nLời kêu gọi hành động: {cta(row, 1)}.",
            f"Headline: Một món nhỏ giúp bạn bớt lăn tăn mỗi ngày.\nSubheadline: {pc} hợp với lúc bạn muốn {comfort}.\nLời kêu gọi hành động: {cta(row, 2)}.",
        ]
    elif row.tone == "professional":
        bodies = [
            f"Headline: {pc} cho nhu cầu sử dụng rõ ràng.\nSubheadline: Phù hợp với {row.target}, nổi bật ở {join_items(row.keywords)}.\nLời kêu gọi hành động: {cta(row, 0)}.",
            f"Headline: Giải pháp gọn cho việc {activity}.\nSubheadline: {pc} giúp hạn chế chuyện {problem} và hỗ trợ sử dụng ổn định hơn.\nLời kêu gọi hành động: {cta(row, 1)}.",
            f"Headline: Chọn {p} khi bạn cần sự tiện dụng thực tế.\nSubheadline: {f1}, {f2} và {f3} là những điểm đáng cân nhắc trước khi đặt hàng.\nLời kêu gọi hành động: {cta(row, 2)}.",
        ]
    elif row.tone == "emotional":
        bodies = [
            f"Headline: {pc} cho những khoảnh khắc bạn muốn thấy yên tâm hơn.\nSubheadline: Một lựa chọn nhỏ giúp {result} trong {profile['daily']}.\nLời kêu gọi hành động: {cta(row, 0)}.",
            f"Headline: Để việc {activity} không còn khiến bạn ngại.\nSubheadline: {pc} làm những chi tiết quen thuộc trở nên dễ chịu hơn nhờ {join_items(row.keywords)}.\nLời kêu gọi hành động: {cta(row, 1)}.",
            f"Headline: Chọn đúng từ những điều nhỏ nhất.\nSubheadline: Khi {problem}, {p} mang lại cảm giác nhẹ lòng hơn cho {row.target}.\nLời kêu gọi hành động: {cta(row, 2)}.",
        ]
    else:
        bodies = [
            f"Headline: {pc} cho việc {activity}.\nSubheadline: Nổi bật với {join_items(row.keywords)}, phù hợp với {row.target}.\nLời kêu gọi hành động: {cta(row, i)}."
            for i in range(3)
        ]
    return versioned(bodies)


def description_benefits(row: RowData, index: int) -> list[str]:
    p = row.product
    profile = get_profile(p)
    activity = profile["activity"]
    f1, f2, f3 = row.keywords[:3]
    if row.tone == "professional":
        sets = [
            [
                f"Phù hợp với {row.target} cần một lựa chọn dễ dùng trong {profile['daily']}.",
                f"Giúp giảm bớt tình trạng {profile['problem']}.",
                f"Hỗ trợ trải nghiệm sử dụng gọn hơn nhờ {join_items(row.keywords)}.",
            ],
            [
                f"Dễ đưa vào thói quen {activity} mà không cần thay đổi quá nhiều.",
                f"Giúp bạn chọn nhanh hơn trước khi đặt hàng.",
                f"Có thêm lựa chọn hỗ trợ từ shop khi cần hỏi kỹ hơn.",
            ],
            [
                f"Tập trung vào nhu cầu thực tế của {row.target}.",
                f"Giúp {profile['result']}.",
                f"Tạo cảm giác mua sắm chắc chắn hơn nhờ thông tin rõ và vừa đủ.",
            ],
        ]
    elif row.tone == "emotional":
        sets = [
            [
                f"Mang lại cảm giác {profile['comfort']}.",
                f"Giúp việc {activity} trở nên nhẹ nhàng hơn.",
                f"Giúp bạn thấy yên tâm hơn khi chọn một món dùng thường xuyên.",
            ],
            [
                f"Giảm bớt nỗi phiền nhỏ như {profile['problem']}.",
                f"Giữ nhịp sống quen thuộc gọn gàng hơn trong {profile['daily']}.",
                f"Tạo cảm giác được chăm chút từ những chi tiết rất đời thường.",
            ],
            [
                f"Phù hợp khi bạn muốn {profile['result']}.",
                f"Biến việc mua sắm thành một lựa chọn dễ chịu, không áp lực.",
                f"Nhắc bạn rằng một món đúng nhu cầu cũng có thể làm ngày nhẹ hơn.",
            ],
        ]
    elif row.tone == "urgent":
        sets = [
            [
                f"Giúp xử lý nhanh chuyện {profile['problem']}.",
                f"Phù hợp khi bạn cần chuẩn bị sớm cho {profile['daily']}.",
                f"Có lời nhắc hành động rõ ràng để không bỏ lỡ ưu đãi đang có.",
            ],
            [
                f"Rút ngắn thời gian chọn mua nhờ các điểm dễ thấy: {f1}, {f2} và {f3}.",
                f"Giúp bạn sẵn sàng hơn cho việc {activity}.",
                f"Thích hợp khi nhu cầu đã rõ và bạn muốn đặt hàng ngay.",
            ],
            [
                f"Đưa {p} vào đúng việc bạn đang cần giải quyết.",
                f"Giúp {profile['result']}.",
                f"Tạo động lực mua sớm mà vẫn dựa trên nhu cầu thật.",
            ],
        ]
    else:
        sets = [
            [
                f"Dễ bắt đầu dùng trong {profile['daily']}.",
                f"Giúp bạn bớt gặp chuyện {profile['problem']}.",
                f"Phù hợp với {row.target} thích lựa chọn đơn giản, rõ ràng.",
            ],
            [
                f"Có {f1} để phục vụ nhu cầu chính.",
                f"Thêm {f2} và {f3} cho trải nghiệm trọn vẹn hơn.",
                f"Dễ hỏi shop khi bạn muốn chọn mẫu hợp với mình.",
            ],
            [
                f"Mang lại cảm giác {profile['comfort']}.",
                f"Giúp {profile['result']}.",
                f"Mua sắm nhẹ nhàng hơn vì các điểm cần xem đã rõ.",
            ],
        ]
    return sets[index]


def description_highlights(row: RowData, index: int) -> list[str]:
    profile = get_profile(row.product)
    activity = profile["activity"]
    f1, f2, f3 = row.keywords[:3]
    extra = row.extra
    sets = [
        [
            f"{cap_first(f1)} hỗ trợ tốt cho nhu cầu {activity}.",
            f"{cap_first(f2)} giúp trải nghiệm dùng hằng ngày linh hoạt hơn.",
            f"{cap_first(f3)} là điểm cộng cho người muốn dùng thường xuyên.",
        ],
        [
            f"Hợp với {row.target} muốn dùng hằng ngày mà không quá cầu kỳ.",
            f"Dễ dùng trong {profile['daily']}.",
            f"{cap_first(extra)}." if extra else f"Dễ kết hợp vào thói quen sử dụng hiện có.",
        ],
        [
            f"Gọn trong cách dùng, rõ trong lợi ích.",
            f"Phù hợp khi bạn ưu tiên sự tiện lợi thay vì những lời quảng cáo lớn.",
            f"Ba điểm đáng nhớ: {f1}, {f2} và {f3}.",
        ],
    ]
    return sets[index]


def build_description(row: RowData) -> str:
    p = row.product
    pc = cap_first(p)
    profile = get_profile(p)
    activity = profile["activity"]
    bonus = extra_phrase(row)
    if row.tone == "professional":
        intros = [
            f"{pc} là lựa chọn thực tế cho {row.target}. Sản phẩm giúp giảm bớt chuyện {profile['problem']} nhờ {join_items(row.keywords)}. {bonus}",
            f"Với {join_items(row.keywords)}, {p} phù hợp cho nhu cầu dùng đều đặn trong {profile['daily']}. Bạn sẽ dễ hình dung cách sản phẩm giúp {profile['result']}.",
            f"Nếu bạn cần một lựa chọn gọn cho việc {activity}, {p} mang lại sự tiện dụng vừa đủ và dễ đưa vào thói quen hằng ngày.",
        ]
    elif row.tone == "emotional":
        intros = [
            f"Có những món nhỏ làm một ngày quen thuộc dễ chịu hơn. {pc} giúp bạn {profile['comfort']} khi {activity}. {bonus}",
            f"Khi {profile['problem']}, một lựa chọn đúng nhu cầu có thể làm bạn nhẹ lòng hơn. {pc} mang đến cảm giác gần gũi nhờ {join_items(row.keywords)}.",
            f"{pc} dành cho những lúc bạn muốn {profile['result']}. Không cần phô trương, sản phẩm ghi điểm bằng các chi tiết dễ dùng mỗi ngày.",
        ]
    elif row.tone == "urgent":
        intros = [
            f"Nếu bạn đang cần {p} cho việc {activity}, đây là lúc nên xem ngay. Sản phẩm có {join_items(row.keywords)} để hỗ trợ nhu cầu sử dụng sớm. {bonus}",
            f"Đừng để {profile['problem']} tiếp tục làm bạn mất thời gian. {pc} giúp {profile['result']} trong {profile['daily']}.",
            f"Nhu cầu đã rõ thì nên chọn sớm. {pc} phù hợp với {row.target} đang muốn một giải pháp gọn, dễ dùng và có hành động mua rõ ràng.",
        ]
    else:
        intros = [
            f"{pc} là món dễ bắt đầu cho {row.target} muốn {profile['result']}. Sản phẩm có {join_items(row.keywords)} và dùng hợp trong {profile['daily']}. {bonus}",
            f"Bạn có thể dùng {p} khi {activity} để bớt gặp chuyện {profile['problem']}. Các chi tiết nhỏ giúp trải nghiệm tự nhiên và dễ chịu hơn.",
            f"Một lựa chọn vừa đủ cho ngày bận rộn: {p} giúp bạn {profile['comfort']} mà không cần chuẩn bị quá cầu kỳ.",
        ]

    bodies: list[str] = []
    for i in range(3):
        benefits = description_benefits(row, i)
        highlights = description_highlights(row, i)
        bodies.append(
            "\n".join(
                [
                    f"Mô tả ngắn: {sentence(intros[i])}",
                    "Lợi ích chính:",
                    f"* {benefits[0]}",
                    f"* {benefits[1]}",
                    f"* {benefits[2]}",
                    "Đặc điểm nổi bật:",
                    f"* {highlights[0]}",
                    f"* {highlights[1]}",
                    f"* {highlights[2]}",
                    f"Lời kêu gọi hành động: {cta(row, i)}.",
                ]
            )
        )
    return versioned(bodies)


def hashtags(row: RowData) -> str:
    base = list(get_profile(row.product)["hashtags"])
    if row.tone == "urgent" and has_offer(row):
        base.append("#UuDaiDangCo")
    unique: list[str] = []
    for tag in base:
        if tag not in unique:
            unique.append(tag)
    return " ".join(unique[:5])


def build_social(row: RowData) -> str:
    p = row.product
    pc = cap_first(p)
    profile = get_profile(p)
    activity = profile["activity"]
    f1, f2, f3 = row.keywords[:3]
    bonus = extra_phrase(row)
    if row.tone == "urgent":
        hooks = [
            f"Đang cần {p}? Xem ngay trước khi bạn lại phải chịu cảnh {profile['problem']}.",
            f"Hôm nay xử lý luôn chuyện {activity} cho gọn.",
            f"{pc} đang đáng xem nếu bạn muốn mua sớm và dùng ngay.",
        ]
        captions = [
            f"{pc} có {join_items(row.keywords)}, phù hợp với {row.target}. {bonus} Nếu đúng thứ bạn đang tìm, hãy nhắn shop ngay để được tư vấn và đặt hàng nhanh hơn.",
            f"Một món nhỏ nhưng dùng đúng lúc thì đỡ phiền thật. Khi {profile['problem']}, {p} giúp {profile['result']}. Đừng để nhu cầu quen thuộc kéo dài thêm.",
            f"Bạn đã biết mình cần {f1}, {f2} và {f3}? Vậy thì {p} là lựa chọn nên xem trong hôm nay, nhất là khi bạn muốn {profile['comfort']}.",
        ]
    else:
        hooks = [
            f"Một món nhỏ cho {profile['daily']} dễ chịu hơn.",
            f"Nếu bạn thường {activity}, {p} đáng để xem thử.",
            f"Gọn hơn một chút, tiện hơn một chút, vậy là đủ vui rồi.",
        ]
        captions = [
            f"{pc} hợp với {row.target} đang muốn chọn một món dễ dùng, dễ hỏi thêm. Có {join_items(row.keywords)}, sản phẩm giúp {profile['result']}. {bonus}",
            f"Mình thích những món giải quyết đúng việc nhỏ hằng ngày. {pc} giúp giảm bớt chuyện {profile['problem']} và dùng khá hợp trong {profile['daily']}.",
            f"Không cần mua sắm quá căng. Nếu bạn cần {f1}, muốn thêm {f2} và {f3}, cứ nhắn shop hỏi mẫu {p} hợp với cách dùng của bạn nhé.",
        ]

    bodies = []
    for i in range(3):
        bodies.append(
            "\n".join(
                [
                    f"Hook: {hooks[i]}",
                    f"Caption: {captions[i]}",
                    f"Lời kêu gọi hành động: {cta(row, i)}.",
                    f"Hashtags: {hashtags(row)}",
                ]
            )
        )
    return versioned(bodies)


def build_email(row: RowData) -> str:
    p = row.product
    pc = cap_first(p)
    profile = get_profile(p)
    activity = profile["activity"]
    f1, f2, f3 = row.keywords[:3]
    bonus = extra_phrase(row)
    if row.tone == "urgent":
        subjects = [f"Xem ngay {p} cho nhu cầu hôm nay", f"Đừng để {profile['problem']} kéo dài", f"{pc}: đặt mua sớm để dùng kịp"]
        previews = ["Một lựa chọn thực tế cho nhu cầu đang có.", "Bạn có thể đặt hàng ngay nếu sản phẩm đúng nhu cầu.", "Thông tin ngắn gọn để chốt nhanh hơn."]
    elif row.tone == "friendly":
        subjects = [f"Gợi ý nhỏ về {p} cho bạn", f"Bạn đang cần {p}? Xem thử nhé", f"Một lựa chọn dễ dùng cho {profile['daily']}"]
        previews = ["Vài điểm dễ hiểu để bạn chọn nhẹ nhàng hơn.", "Nếu hợp nhu cầu, shop có thể gợi ý thêm.", "Một món nhỏ nhưng dùng đúng lúc sẽ tiện hơn nhiều."]
    elif row.tone == "emotional":
        subjects = [f"Một lựa chọn giúp bạn yên tâm hơn", f"Khi {p} hợp với điều bạn đang cần", f"Cho {profile['daily']} dễ chịu hơn một chút"]
        previews = ["Những chi tiết nhỏ có thể làm ngày của bạn nhẹ hơn.", "Chọn đúng món đôi khi là chọn cảm giác an tâm.", "Một gợi ý vừa phải cho nhu cầu quen thuộc."]
    else:
        subjects = [f"Gợi ý {p} cho {row.target}", f"{pc} cho nhu cầu sử dụng hằng ngày", f"Bạn có thể cân nhắc {p}"]
        previews = [f"Nổi bật với {join_items(row.keywords)}.", f"Phù hợp cho việc {activity}.", "Một lựa chọn thực tế, dễ hỏi thêm trước khi mua."]

    bodies = [
        f"Nếu bạn thường {activity}, {p} có thể giúp giảm bớt chuyện {profile['problem']}. Sản phẩm có {join_items(row.keywords)}, phù hợp với {row.target}. {bonus} Bạn có thể xem chi tiết hoặc nhắn shop để chọn đúng mẫu.",
        f"{pc} được làm cho những nhu cầu rất đời thường: {profile['daily']} cần tiện hơn, gọn hơn và bớt phiền hơn. Với {f1}, {f2} và {f3}, sản phẩm giúp {profile['result']}.",
        f"Trước khi đặt hàng, bạn chỉ cần nghĩ xem mình sẽ dùng {p} trong tình huống nào. Nếu câu trả lời là {activity}, lựa chọn này đáng để xem kỹ. Shop có thể tư vấn thêm để bạn chọn đúng nhu cầu.",
    ]

    parts = []
    for i in range(3):
        parts.append(
            "\n".join(
                [
                    f"Subject: {subjects[i]}",
                    f"Preview text: {previews[i]}",
                    "Lời chào: Chào bạn,",
                    f"Nội dung chính: {bodies[i]}",
                    f"Lời kêu gọi hành động: {cta(row, i)}.",
                ]
            )
        )
    return versioned(parts)


def build_cta(row: RowData) -> str:
    p = row.product
    profile = get_profile(p)
    activity = profile["activity"]
    f1, f2, f3 = row.keywords[:3]
    if row.tone == "urgent":
        micros = [
            f"Có {join_items(row.keywords)}; đặt sớm nếu bạn đang cần dùng trong {profile['daily']}.",
            f"Nhắn shop ngay để được tư vấn mẫu phù hợp với việc {activity}.",
            f"{extra_phrase(row) or sentence('Shop sẵn sàng hỗ trợ bạn chọn nhanh hơn')}",
        ]
    elif row.tone == "professional":
        micros = [
            f"Xem {f1}, {f2} và {f3} trước khi quyết định.",
            f"Shop hỗ trợ tư vấn theo nhu cầu của {row.target}.",
            f"Phù hợp khi bạn cần một lựa chọn thực tế cho việc {activity}.",
        ]
    elif row.tone == "emotional":
        micros = [
            f"Một bước nhỏ để {profile['daily']} dễ chịu hơn.",
            f"Hỏi thêm nếu bạn muốn chọn {p} thật vừa với nhu cầu.",
            f"Dành cho lúc bạn muốn {profile['comfort']}.",
        ]
    else:
        micros = [
            f"Có {join_items(row.keywords)}; xem thử có hợp với bạn không nhé.",
            f"Kể shop nghe bạn dùng trong tình huống nào, shop gợi ý mẫu phù hợp hơn.",
            f"Một lựa chọn gọn cho việc {activity}.",
        ]
    return versioned(
        [
            f"Lời kêu gọi hành động chính: {cta(row, 0)}\nMicrocopy: {micros[0]}",
            f"Lời kêu gọi hành động chính: {cta(row, 1)}\nMicrocopy: {micros[1]}",
            f"Lời kêu gọi hành động chính: {cta(row, 2)}\nMicrocopy: {micros[2]}",
        ]
    )


def build_output(row: RowData) -> str:
    builders = {
        "headline": build_headline,
        "description": build_description,
        "social": build_social,
        "email": build_email,
        "cta": build_cta,
    }
    if row.content_type not in builders:
        raise ValueError(f"Dòng {row.row_number}: chưa hỗ trợ type {row.content_type}")
    return builders[row.content_type](row)


def split_versions(output: str) -> list[tuple[int, str]]:
    pattern = re.compile(r"^Phiên bản (\d+):\s*$", re.MULTILINE)
    matches = list(pattern.finditer(output.strip()))
    versions = []
    for i, match in enumerate(matches):
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(output)
        versions.append((int(match.group(1)), output[start:end].strip()))
    return versions


def line_label(line: str) -> str | None:
    if line.startswith("* "):
        return None
    match = re.match(r"^([^:]{1,80}):", line)
    return match.group(1).strip() if match else None


def validate_format(row: RowData, output: str) -> list[str]:
    errors: list[str] = []
    expected_count = TYPE_VERSION_COUNTS[row.content_type]
    expected_numbers = list(range(1, expected_count + 1))
    versions = split_versions(output)
    if [number for number, _ in versions] != expected_numbers:
        return [f"Dòng {row.row_number}: sai số phiên bản"]
    if len({norm_compare(segment) for _, segment in versions}) != len(versions):
        errors.append(f"Dòng {row.row_number}: có phiên bản trùng trong cùng ô output")
    allowed = TYPE_FIELDS[row.content_type]
    for number, segment in versions:
        lines = [line for line in segment.split("\n") if line.strip()]
        labels = [label for label in (line_label(line) for line in lines) if label]
        if labels != allowed:
            errors.append(f"Dòng {row.row_number} phiên bản {number}: field sai {labels}, cần {allowed}")
        bullet_count = sum(1 for line in lines if line.startswith("* "))
        if row.content_type == "description" and bullet_count != 6:
            errors.append(f"Dòng {row.row_number} phiên bản {number}: description cần 6 bullet, hiện có {bullet_count}")
        if row.content_type == "social":
            tag_lines = [line for line in lines if line.startswith("Hashtags:")]
            if len(tag_lines) != 1:
                errors.append(f"Dòng {row.row_number} phiên bản {number}: thiếu Hashtags")
            else:
                count = len(re.findall(r"#[\wÀ-ỹ]+", tag_lines[0], re.UNICODE))
                if count < 3 or count > 6:
                    errors.append(f"Dòng {row.row_number} phiên bản {number}: hashtag count {count}")
    return errors


def validate_outputs(rows: list[RowData], outputs: dict[int, str]) -> list[str]:
    errors: list[str] = []
    seen: dict[str, int] = {}
    groups: dict[tuple[str, str], list[RowData]] = defaultdict(list)
    for row in rows:
        output = outputs[row.row_number]
        lowered = output.lower()
        for phrase in META_PHRASES + FORBIDDEN_GENERAL:
            if phrase.lower() in lowered:
                errors.append(f"Dòng {row.row_number}: còn cụm cấm/meta '{phrase}'")
        if row.tone == "urgent":
            for phrase in URGENT_WEAK:
                if phrase in lowered:
                    errors.append(f"Dòng {row.row_number}: urgent tone còn câu yếu '{phrase}'")
        normalized = norm_compare(output)
        if normalized in seen:
            errors.append(f"Dòng {row.row_number}: output trùng dòng {seen[normalized]}")
        else:
            seen[normalized] = row.row_number
        errors.extend(validate_format(row, output))
        groups[(norm_compare(row.product), row.content_type)].append(row)

    for (_product, _content_type), group_rows in groups.items():
        for i, left in enumerate(group_rows):
            for right in group_rows[i + 1 :]:
                if left.tone == right.tone:
                    continue
                left_output = norm_compare(outputs[left.row_number])
                right_output = norm_compare(outputs[right.row_number])
                ratio = SequenceMatcher(None, left_output, right_output).ratio()
                if ratio >= 0.90:
                    errors.append(f"Dòng {left.row_number} và {right.row_number}: tone khác nhưng output quá giống ({ratio:.2f})")
    return errors


def main() -> int:
    input_path = find_input_path()
    output_path = input_path.with_name(OUTPUT_FILE)
    wb = load_workbook(input_path)
    ws, header_map = find_target_sheet(wb)
    rows = read_rows(ws, header_map)
    outputs = {row.row_number: build_output(row) for row in rows}
    errors = validate_outputs(rows, outputs)
    if errors:
        print("Validation: FAIL")
        for error in errors[:80]:
            print("-", error)
        return 1

    output_column = header_map["output"]
    for row in rows:
        cell = ws.cell(row=row.row_number, column=output_column)
        cell.value = outputs[row.row_number]
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)

    reread = load_workbook(output_path, read_only=True)
    final_ws, final_header_map = find_target_sheet(reread)
    final_rows = read_rows(final_ws, final_header_map)
    final_outputs = {
        row.row_number: normalize_text(final_ws.cell(row=row.row_number, column=final_header_map["output"]).value)
        for row in final_rows
    }
    final_errors = validate_outputs(final_rows, final_outputs)

    duplicate_rows = sum(count for count in Counter(norm_compare(value) for value in final_outputs.values()).values() if count > 1)
    print(f"Final file: {output_path.resolve()}")
    print(f"Rows processed: {len(final_rows)}")
    print(f"Duplicate outputs after final: {duplicate_rows}")
    print(f"Validation: {'PASS' if not final_errors else 'FAIL'}")
    if final_errors:
        for error in final_errors[:80]:
            print("-", error)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
