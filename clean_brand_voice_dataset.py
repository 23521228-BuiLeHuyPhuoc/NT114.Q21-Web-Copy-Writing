# -*- coding: utf-8 -*-
"""Clean Vietnamese brand-voice fine-tuning examples.

The script intentionally uses openpyxl for XLSX I/O and does not call any
external APIs. It keeps row order, writes a new workbook, and generates a
Markdown report with validation results.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
import re
import sys
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


SOURCE_FILE = "01_train_examples_brand_voice_UTF8_FIXED.xlsx"
CLEANED_FILE = "01_train_examples_brand_voice_UTF8_CLEANED.xlsx"
REPORT_FILE = "01_train_examples_brand_voice_UTF8_CLEANING_REPORT.md"

REQUIRED_COLUMNS = ["input", "output", "industry", "tone", "type", "product"]

VALID_INDUSTRIES = {
    "ecommerce",
    "realestate",
    "technology",
    "fnb",
    "healthcare",
    "education",
    "finance",
    "fashion",
    "business",
    "travel",
}
VALID_TYPES = {"headline", "description", "social", "email", "cta", "landing", "seo", "review"}
VALID_TONES = {"urgent", "professional", "friendly", "luxury", "humorous", "emotional"}

FORBIDDEN_INPUT_PHRASES = [
    "Temperature tham khảo",
    "Giới hạn output tối đa",
    "maxOutputTokens",
    "modelMode",
    "model",
    "fineTunedModelId",
    "templateId",
    "projectId",
    "Model AI được chọn",
]

FORBIDDEN_OUTPUT_PHRASES = [
    "Góc cần biết",
    "Lưu ý",
    "Ghi chú",
    "Dưới đây là",
    "Tôi đã sửa",
]

FORBIDDEN_CLAIMS = [
    "cam kết 100%",
    "tốt nhất thị trường",
    "chữa khỏi",
    "đảm bảo lợi nhuận",
    "duy nhất hôm nay",
    "sắp hết hàng",
]

URGENT_WEAK_PHRASES = [
    "không cần mua vội",
    "lưu lại để xem sau",
    "cân nhắc từ từ",
]

TYPE_VERSION_COUNTS = {
    "headline": 3,
    "description": 3,
    "social": 3,
    "email": 3,
    "cta": 3,
    "landing": 1,
    "seo": 1,
    "review": 1,
}

TYPE_FIELDS = {
    "headline": ["Headline", "Subheadline", "Lời kêu gọi hành động"],
    "description": ["Mô tả ngắn", "Lợi ích chính", "Đặc điểm nổi bật", "Lời kêu gọi hành động"],
    "social": ["Hook", "Caption", "Lời kêu gọi hành động", "Hashtags"],
    "email": ["Subject", "Preview text", "Lời chào", "Nội dung chính", "Lời kêu gọi hành động"],
    "cta": ["Lời kêu gọi hành động chính", "Microcopy"],
    "landing": [
        "Hero headline",
        "Hero subheadline",
        "Lợi ích chính",
        "Nội dung chính",
        "Bằng chứng thuyết phục",
        "Lời kêu gọi hành động",
    ],
    "seo": ["SEO title", "Meta description", "H1", "Mở bài", "Dàn ý nội dung", "Lời kêu gọi hành động"],
    "review": [
        "Tên người đánh giá",
        "Bối cảnh sử dụng",
        "Nội dung review",
        "Điểm nổi bật được nhắc đến",
        "Kết luận / Gợi ý",
    ],
}

ALL_FIELD_LABELS = {field for fields in TYPE_FIELDS.values() for field in fields}

TYPE_NAMES = {
    "headline": "tiêu đề quảng cáo",
    "description": "mô tả sản phẩm",
    "social": "bài đăng mạng xã hội",
    "email": "email marketing",
    "cta": "lời kêu gọi hành động",
    "landing": "landing page",
    "seo": "nội dung SEO",
    "review": "review khách hàng",
}

TONE_NAMES = {
    "urgent": "khẩn cấp",
    "professional": "chuyên nghiệp",
    "friendly": "thân thiện",
    "luxury": "cao cấp",
    "humorous": "hài hước",
    "emotional": "cảm xúc",
}


@dataclass
class ExampleRow:
    excel_row: int
    input: str
    output: str
    industry: str
    tone: str
    type: str
    product: str
    keywords: list[str]
    target: str
    extra: str
    desired_length: str


def normalize_ws_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFC", text)
    return text.replace("\r\n", "\n").replace("\r", "\n")


def normalize_key(value: object) -> str:
    return normalize_ws_text(value).strip().lower()


def normalize_for_compare(value: str) -> str:
    text = unicodedata.normalize("NFC", value).lower()
    text = re.sub(r"\s+", " ", text).strip()
    return text


def sentence(text: str) -> str:
    text = re.sub(r"\s+", " ", text.strip()).strip(" .;,")
    if not text:
        return ""
    if text[-1] not in ".!?":
        text += "."
    return text


def cap_first(text: str) -> str:
    text = text.strip()
    if not text:
        return text
    return text[0].upper() + text[1:]


def strip_final_dot(text: str) -> str:
    return re.sub(r"[\s.]+$", "", text.strip())


def join_items(items: list[str]) -> str:
    clean = [strip_final_dot(item) for item in items if strip_final_dot(item)]
    if not clean:
        return "thông tin sản phẩm rõ ràng"
    if len(clean) == 1:
        return clean[0]
    if len(clean) == 2:
        return f"{clean[0]} và {clean[1]}"
    return f"{', '.join(clean[:-1])} và {clean[-1]}"


def safe_value(row_values: tuple[object, ...], index: int) -> str:
    if index >= len(row_values):
        return ""
    return normalize_ws_text(row_values[index]).strip()


def resolve_source_path() -> Path:
    root = Path.cwd()
    preferred = [
        root / SOURCE_FILE,
        root / "fine_tuning_excel_pack_brand_voice_UTF8_FIXED" / SOURCE_FILE,
        root / "fine_tuning_excel_pack_brand_voice" / SOURCE_FILE,
    ]
    for candidate in preferred:
        if candidate.exists():
            return candidate

    matches = sorted(
        path for path in root.rglob(SOURCE_FILE) if CLEANED_FILE not in path.name and ".git" not in path.parts
    )
    if not matches:
        raise FileNotFoundError(f"Không tìm thấy file nguồn: {SOURCE_FILE}")
    return matches[0]


def find_target_sheet(workbook) -> tuple[object, dict[str, int]]:
    for worksheet in workbook.worksheets:
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        header_map = {normalize_key(value): idx for idx, value in enumerate(header_row) if value is not None}
        if all(column in header_map for column in REQUIRED_COLUMNS):
            return worksheet, header_map
    raise ValueError("Không tìm thấy sheet có đủ cột: input | output | industry | tone | type | product")


def validate_required_columns(header_map: dict[str, int]) -> list[str]:
    missing = [column for column in REQUIRED_COLUMNS if column not in header_map]
    if missing:
        return ["Thiếu cột bắt buộc: " + ", ".join(missing)]
    return []


def clean_input_text(input_text: str) -> str:
    cleaned_lines: list[str] = []
    for line in normalize_ws_text(input_text).split("\n"):
        lowered = line.lower()
        if any(phrase.lower() in lowered for phrase in FORBIDDEN_INPUT_PHRASES):
            continue
        cleaned_lines.append(line.rstrip())

    while cleaned_lines and not cleaned_lines[0].strip():
        cleaned_lines.pop(0)
    while cleaned_lines and not cleaned_lines[-1].strip():
        cleaned_lines.pop()
    return "\n".join(cleaned_lines).strip()


def extract_label(input_text: str, label: str) -> str:
    wanted = label.lower()
    for raw_line in normalize_ws_text(input_text).split("\n"):
        line = raw_line.strip()
        if line.lower().startswith(wanted + ":"):
            return strip_final_dot(line.split(":", 1)[1])
    return ""


def split_keywords(value: str) -> list[str]:
    parts = [strip_final_dot(part) for part in re.split(r"[;,\n]+", value) if strip_final_dot(part)]
    return parts[:5]


def pad_keywords(keywords: list[str], product: str) -> list[str]:
    padded = [item for item in keywords if item]
    fallbacks = [
        f"phù hợp với nhu cầu của {product}",
        "thông tin rõ ràng để dễ so sánh",
        "trải nghiệm mua hàng dễ cân nhắc",
    ]
    for fallback in fallbacks:
        if len(padded) >= 3:
            break
        padded.append(fallback)
    return padded[:3]


def read_examples(worksheet, header_map: dict[str, int]) -> list[ExampleRow]:
    examples: list[ExampleRow] = []
    for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        original_input = safe_value(values, header_map["input"])
        product = safe_value(values, header_map["product"])
        clean_input = clean_input_text(original_input)
        keywords_text = extract_label(original_input, "Từ khóa chính")
        target = extract_label(original_input, "Đối tượng mục tiêu")
        extra = extract_label(original_input, "Thông tin bổ sung")
        desired_length = extract_label(original_input, "Độ dài mong muốn")

        examples.append(
            ExampleRow(
                excel_row=excel_row,
                input=clean_input,
                output=safe_value(values, header_map["output"]),
                industry=safe_value(values, header_map["industry"]).lower(),
                tone=safe_value(values, header_map["tone"]).lower(),
                type=safe_value(values, header_map["type"]).lower(),
                product=product,
                keywords=pad_keywords(split_keywords(keywords_text), product),
                target=target or "người đang cân nhắc sản phẩm này",
                extra=extra,
                desired_length=desired_length,
            )
        )
    return examples


def has_offer(extra: str) -> bool:
    lowered = extra.lower()
    offer_markers = ["freeship", "giảm", "tặng", "ưu đãi", "miễn phí", "đổi", "voucher"]
    return any(marker in lowered for marker in offer_markers)


def extra_sentence(example: ExampleRow) -> str:
    if not example.extra:
        return ""
    return sentence(f"Thông tin đi kèm là {example.extra}")


def cta_for(example: ExampleRow, variant: int = 0) -> str:
    tone = example.tone
    product = example.product
    urgent_ctas = [
        "Xem ưu đãi ngay" if has_offer(example.extra) else "Xem chi tiết ngay",
        "Đặt mua hôm nay",
        "Nhắn shop để giữ ưu đãi" if has_offer(example.extra) else "Nhắn shop để được tư vấn ngay",
    ]
    ctas = {
        "urgent": urgent_ctas,
        "professional": ["Xem thông tin sản phẩm", "Yêu cầu tư vấn chi tiết", "Đối chiếu thông số ngay"],
        "friendly": ["Xem thử sản phẩm", "Nhắn shop để được gợi ý", "Chọn mẫu hợp với bạn"],
        "luxury": ["Khám phá lựa chọn tinh tế", "Đặt lịch tư vấn riêng", "Chọn trải nghiệm xứng tầm"],
        "humorous": ["Xem ngay cho khỏi tò mò", "Nhắn shop hỏi nhanh", "Chọn món hợp gu"],
        "emotional": ["Chọn trải nghiệm dễ chịu hơn", "Nhắn shop để được gợi ý", "Xem lựa chọn phù hợp với bạn"],
    }
    selected = ctas.get(tone, ctas["friendly"])[variant % 3]
    if tone in {"professional", "friendly", "emotional"} and variant == 2:
        return selected
    if tone == "humorous" and variant == 2:
        return selected
    if tone == "luxury" and variant == 0:
        return selected
    if product and tone == "professional" and variant == 0:
        return selected
    return selected


def tone_intro(example: ExampleRow, variant: int) -> str:
    product = example.product
    audience = example.target
    features = join_items(example.keywords)
    intros = {
        "urgent": [
            f"Khi nhu cầu đã rõ, {product} là lựa chọn nên được xem sớm để không bỏ lỡ thông tin đang có",
            f"Nếu {audience} đang cần một lựa chọn thực tế, {product} giúp bạn ra quyết định nhanh hơn dựa trên {features}",
            f"Đây là thời điểm phù hợp để kiểm tra {product} trước khi nhu cầu hằng ngày tiếp tục bị trì hoãn",
        ],
        "professional": [
            f"{cap_first(product)} được trình bày như một lựa chọn rõ tiêu chí cho {audience}",
            f"Nội dung tập trung vào các điểm có thể đánh giá trực tiếp của {product}",
            f"Với {features}, {product} phù hợp để đưa vào danh sách so sánh có cơ sở",
        ],
        "friendly": [
            f"{cap_first(product)} là lựa chọn dễ bắt đầu cho {audience}",
            f"Nếu bạn muốn mua sắm nhẹ nhàng hơn, {product} có các điểm đáng xem như {features}",
            f"Một món phù hợp đôi khi chỉ cần đúng nhu cầu và dễ dùng mỗi ngày như {product}",
        ],
        "luxury": [
            f"{cap_first(product)} hướng đến trải nghiệm chọn lọc, chỉn chu và vừa vặn với nhu cầu sử dụng",
            f"Với {audience}, {product} tạo ấn tượng bằng sự tinh tế trong từng điểm đáng chú ý",
            f"Không cần phô trương, {product} thuyết phục bằng chất lượng cảm nhận và thông tin rõ ràng",
        ],
        "humorous": [
            f"{cap_first(product)} dành cho lúc bạn muốn giải quyết nhu cầu thật mà không biến việc mua sắm thành một cuộc họp dài",
            f"Nếu {audience} đang phân vân, {product} có đủ điểm để bạn gật gù nhanh hơn",
            f"Mua đúng món thì vui nhẹ cả ngày, nhất là khi {product} khớp với điều bạn đang cần",
        ],
        "emotional": [
            f"Có những chi tiết nhỏ làm một ngày quen thuộc trở nên dễ chịu hơn, và {product} được viết cho cảm giác đó",
            f"Với {audience}, {product} không chỉ là món để mua mà còn là cách chăm chút cho nhu cầu hằng ngày",
            f"Khi bạn muốn thấy yên tâm hơn trong lựa chọn của mình, {product} mang đến một điểm bắt đầu vừa phải",
        ],
    }
    return intros.get(example.tone, intros["friendly"])[variant % 3]


def hashtag_token(text: str, max_words: int = 4) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    words = re.findall(r"[A-Za-z0-9]+", text)
    if not words:
        return "MarketingViet"
    return "".join(word[:1].upper() + word[1:] for word in words[:max_words])


def make_hashtags(example: ExampleRow) -> str:
    tokens = [
        hashtag_token(example.product),
        hashtag_token(example.keywords[0]),
        hashtag_token(example.target),
        hashtag_token(TYPE_NAMES.get(example.type, example.type)),
    ]
    unique: list[str] = []
    for token in tokens:
        if token and token not in unique:
            unique.append(token)
    while len(unique) < 3:
        unique.append(f"MuaSam{len(unique) + 1}")
    return " ".join(f"#{token}" for token in unique[:6])


def versioned(bodies: list[str]) -> str:
    return "\n\n".join(f"Phiên bản {idx}:\n{body.strip()}" for idx, body in enumerate(bodies, start=1))


def build_headline_output(example: ExampleRow) -> str:
    product = example.product
    product_cap = cap_first(product)
    audience = example.target
    features = join_items(example.keywords)
    f1, f2, f3 = example.keywords[:3]
    extra = extra_sentence(example)

    if example.tone == "urgent":
        urgent_headline = (
            f"Cần {product} đúng nhu cầu? Xem ngay trước khi bỏ lỡ ưu đãi phù hợp."
            if has_offer(example.extra)
            else f"Cần {product} đúng nhu cầu? Xem ngay trước khi bỏ lỡ lựa chọn phù hợp."
        )
        bodies = [
            "\n".join(
                [
                    f"Headline: {urgent_headline}",
                    f"Subheadline: Dành cho {audience}, tập trung vào {features}. {extra}",
                    f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: {product_cap} giúp bạn xử lý nhu cầu {f1} ngay hôm nay.",
                    f"Subheadline: Ba điểm cần xem nhanh là {f1}, {f2} và {f3}; đủ rõ để quyết định sớm hơn.",
                    f"Lời kêu gọi hành động: {cta_for(example, 1)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Đến lúc chọn {product} cho việc bạn đang cần mỗi ngày.",
                    f"Subheadline: Phù hợp với {audience} và có thông tin cụ thể để bạn chốt nhanh mà vẫn kiểm soát lựa chọn.",
                    f"Lời kêu gọi hành động: {cta_for(example, 2)}.",
                ]
            ),
        ]
    elif example.tone == "friendly":
        bodies = [
            "\n".join(
                [
                    f"Headline: {product_cap} cho ngày dùng tiện hơn.",
                    f"Subheadline: Phù hợp với {audience}, dễ xem nhờ các điểm như {features}.",
                    f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Chọn {product} theo cách nhẹ nhàng và dễ hiểu.",
                    f"Subheadline: Bạn có thể bắt đầu từ {f1}, rồi xem thêm {f2} và {f3} trước khi hỏi shop.",
                    f"Lời kêu gọi hành động: {cta_for(example, 1)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Một lựa chọn gần gũi cho {audience}.",
                    f"Subheadline: {product_cap} tập trung vào nhu cầu thật, thông tin rõ và không tạo cảm giác bị thúc ép.",
                    f"Lời kêu gọi hành động: {cta_for(example, 2)}.",
                ]
            ),
        ]
    elif example.tone == "professional":
        bodies = [
            "\n".join(
                [
                    f"Headline: {product_cap} cho nhu cầu sử dụng rõ tiêu chí.",
                    f"Subheadline: Nội dung tập trung vào {features}, phù hợp để {audience} đánh giá trước khi mua.",
                    f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Đánh giá {product} qua những điểm có thể kiểm chứng.",
                    f"Subheadline: {f1}, {f2} và {f3} là các thông tin chính giúp việc so sánh trở nên gọn hơn.",
                    f"Lời kêu gọi hành động: {cta_for(example, 1)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Một lựa chọn thực tế cho {audience}.",
                    f"Subheadline: {product_cap} được giới thiệu bằng thông tin rõ ràng, không phóng đại lợi ích.",
                    f"Lời kêu gọi hành động: {cta_for(example, 2)}.",
                ]
            ),
        ]
    elif example.tone == "luxury":
        bodies = [
            "\n".join(
                [
                    f"Headline: {product_cap} cho trải nghiệm sử dụng tinh tế hơn.",
                    f"Subheadline: Lựa chọn dành cho {audience}, nổi bật bằng {features} trong một cách tiếp cận chỉn chu.",
                    f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Chọn {product} khi từng chi tiết đều cần vừa vặn.",
                    f"Subheadline: {f1}, {f2} và {f3} tạo nên cảm giác được cân nhắc kỹ trước khi đưa vào sử dụng.",
                    f"Lời kêu gọi hành động: {cta_for(example, 1)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Một lựa chọn có gu cho {audience}.",
                    f"Subheadline: {product_cap} đặt trọng tâm vào trải nghiệm, chất lượng cảm nhận và sự gọn gàng trong quyết định.",
                    f"Lời kêu gọi hành động: {cta_for(example, 2)}.",
                ]
            ),
        ]
    elif example.tone == "humorous":
        bodies = [
            "\n".join(
                [
                    f"Headline: {product_cap} giúp việc đang cần bớt làm bạn nhăn trán.",
                    f"Subheadline: Có {features}, nên {audience} có thêm lý do để xem kỹ mà không cần đọc một bài dài.",
                    f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Chọn {product} cho nhu cầu thật, bớt vòng vo một chút.",
                    f"Subheadline: {f1}, {f2} và {f3} là ba điểm đủ để bắt đầu cuộc cân nhắc nghiêm túc nhưng nhẹ nhàng.",
                    f"Lời kêu gọi hành động: {cta_for(example, 1)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Món này có thể hợp với bạn hơn là chiếc giỏ hàng đang chờ.",
                    f"Subheadline: {product_cap} hướng đến {audience}, nói thẳng vào nhu cầu thay vì làm quá câu chuyện.",
                    f"Lời kêu gọi hành động: {cta_for(example, 2)}.",
                ]
            ),
        ]
    elif example.tone == "emotional":
        bodies = [
            "\n".join(
                [
                    f"Headline: {product_cap} cho những lúc bạn muốn thấy yên tâm hơn.",
                    f"Subheadline: Dành cho {audience}, sản phẩm tập trung vào {features} để nhu cầu hằng ngày nhẹ đi một chút.",
                    f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Một lựa chọn nhỏ có thể làm ngày của bạn dễ chịu hơn.",
                    f"Subheadline: {product_cap} không nói quá, chỉ đặt {f1}, {f2} và {f3} vào đúng điều bạn đang cần.",
                    f"Lời kêu gọi hành động: {cta_for(example, 1)}.",
                ]
            ),
            "\n".join(
                [
                    f"Headline: Khi chọn đúng {product}, cảm giác an tâm đến từ những chi tiết vừa đủ.",
                    f"Subheadline: Phù hợp với {audience} đang muốn một quyết định mua hàng bớt phân vân hơn.",
                    f"Lời kêu gọi hành động: {cta_for(example, 2)}.",
                ]
            ),
        ]
    else:
        bodies = [
            "\n".join(
                [
                    f"Headline: {product_cap} cho nhu cầu sử dụng hằng ngày.",
                    f"Subheadline: Phù hợp với {audience}, tập trung vào {features}.",
                    f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
                ]
            )
            for _ in range(3)
        ]
    return versioned(bodies)


def benefit_lines(example: ExampleRow, variant: int) -> list[str]:
    product = example.product
    audience = example.target
    f1, f2, f3 = example.keywords[:3]
    tone = example.tone
    if tone == "urgent":
        sets = [
            [
                f"Giúp {audience} xử lý nhu cầu {f1} sớm hơn.",
                f"Làm rõ điểm {f2} để bạn không mất thời gian so sánh lan man.",
                f"Đưa {f3} vào quyết định mua ngay khi thông tin còn mới.",
            ],
            [
                f"Tập trung vào nhu cầu đang diễn ra của {audience}.",
                f"Giúp bạn xem nhanh ba tiêu chí chính trước khi đặt hàng.",
                f"Phù hợp khi muốn chốt lựa chọn trong hôm nay mà vẫn dựa trên thông tin cụ thể.",
            ],
            [
                f"Rút ngắn bước cân nhắc bằng các điểm chính của {product}.",
                f"Hỗ trợ quyết định nhanh hơn nhờ {f1} và {f2}.",
                f"Giữ lời kêu gọi hành động rõ ràng, phù hợp với nhu cầu mua sớm.",
            ],
        ]
    elif tone == "professional":
        sets = [
            [
                f"Cung cấp tiêu chí {f1} để đánh giá sản phẩm rõ ràng hơn.",
                f"Làm nổi bật {f2} như một thông tin cần đối chiếu trước khi mua.",
                f"Giúp {audience} hiểu vai trò của {f3} trong quá trình sử dụng.",
            ],
            [
                f"Tập trung vào lợi ích thực tế thay vì cảm xúc quá mức.",
                f"Phù hợp để so sánh {product} với các lựa chọn cùng nhu cầu.",
                f"Giảm mơ hồ trong quyết định nhờ các điểm thông tin cụ thể.",
            ],
            [
                f"Trình bày rõ nhóm người dùng phù hợp là {audience}.",
                f"Sắp xếp thông tin theo lợi ích, đặc điểm và hành động tiếp theo.",
                f"Giữ giọng văn gọn, đáng tin và dễ đưa vào tài liệu bán hàng.",
            ],
        ]
    elif tone == "luxury":
        sets = [
            [
                f"Mang lại cảm giác chọn lọc hơn khi {audience} cân nhắc {product}.",
                f"Nhấn vào {f1} như một phần của trải nghiệm sử dụng chỉn chu.",
                f"Kết hợp {f2} và {f3} để tạo sự yên tâm tinh tế.",
            ],
            [
                f"Giữ thông điệp trang nhã, không dùng cách bán hàng quá ồn ào.",
                f"Tôn lên chất lượng cảm nhận từ những điểm sản phẩm đã có.",
                f"Phù hợp với người mua cần sự vừa vặn và kỹ lưỡng.",
            ],
            [
                f"Đưa {product} vào bối cảnh sử dụng có chọn lọc.",
                f"Làm rõ giá trị qua chi tiết, không cần phóng đại.",
                f"Tạo cảm giác tin cậy bằng nhịp câu chậm và từ ngữ tinh tế.",
            ],
        ]
    elif tone == "humorous":
        sets = [
            [
                f"Giúp {audience} hiểu nhanh sản phẩm mà không thấy khô như bảng thông số.",
                f"Biến {f1} thành lý do dễ nhớ khi cân nhắc.",
                f"Nhắc đến {f2} và {f3} theo cách nhẹ nhàng nhưng vẫn rõ ý.",
            ],
            [
                f"Làm thông tin sản phẩm bớt cứng mà không làm giảm độ tin cậy.",
                f"Giữ nội dung ngắn gọn cho người đang lướt nhanh.",
                f"Tạo cảm giác vui nhẹ trước khi người đọc chuyển sang hành động.",
            ],
            [
                f"Nói đúng nhu cầu của {audience} bằng một nhịp câu dễ nhớ.",
                f"Tập trung vào các điểm thật của {product}, không thêm lời hứa quá đà.",
                f"Khuyến khích hỏi thêm mà không ép người đọc quá mạnh.",
            ],
        ]
    elif tone == "emotional":
        sets = [
            [
                f"Giúp {audience} thấy yên tâm hơn khi lựa chọn {product}.",
                f"Gắn {f1} với cảm giác thuận tiện trong tình huống đời thực.",
                f"Làm {f2} và {f3} trở thành những chi tiết nhỏ nhưng đáng nhớ.",
            ],
            [
                f"Nhấn vào cảm giác nhẹ nhõm khi nhu cầu hằng ngày được xử lý tốt hơn.",
                f"Tạo kết nối gần gũi mà không làm nội dung trở nên bi lụy.",
                f"Giữ thông điệp mua hàng mềm mại nhưng vẫn có hành động rõ.",
            ],
            [
                f"Đặt {product} vào khoảnh khắc người dùng thật sự cần sự tiện lợi.",
                f"Cho người đọc thấy lựa chọn này có thể làm ngày của họ dễ chịu hơn.",
                f"Tập trung vào sự an tâm thay vì lời quảng cáo quá lớn.",
            ],
        ]
    else:
        sets = [
            [
                f"Giúp bạn xem nhanh {f1} trước khi mua.",
                f"Làm rõ {f2} trong quá trình sử dụng.",
                f"Bổ sung {f3} để dễ cân nhắc hơn.",
            ],
            [
                f"Phù hợp với {audience}.",
                f"Thông tin trình bày dễ hiểu.",
                f"Có lời kêu gọi hành động rõ ràng.",
            ],
            [
                f"Tập trung vào nhu cầu thực tế.",
                f"Dễ đưa vào quyết định mua hàng.",
                f"Không dùng lời hứa quá đà.",
            ],
        ]
    return sets[variant % 3]


def highlight_lines(example: ExampleRow, variant: int) -> list[str]:
    f1, f2, f3 = example.keywords[:3]
    extra = example.extra
    base = [
        f"Điểm nổi bật đầu tiên là {f1}.",
        f"Điểm tiếp theo cần nhắc đến là {f2}.",
        f"Điểm bổ sung giúp hoàn thiện lựa chọn là {f3}.",
    ]
    if variant == 1:
        base = [
            f"Sản phẩm được mô tả trực tiếp qua {f1}.",
            f"Thông tin {f2} giúp người mua có thêm cơ sở đánh giá.",
            f"Yếu tố {f3} làm nội dung bớt chung chung.",
        ]
    elif variant == 2:
        base = [
            f"Chỉ nhắc đến các điểm đã có, không thêm tính năng mới.",
            f"Các điểm được nhấn mạnh gồm {join_items(example.keywords)}.",
            f"Thông điệp hướng về {example.target}.",
        ]
    if extra and variant == 0:
        base[2] = f"Thông tin đi kèm là {extra}."
    return base


def build_description_output(example: ExampleRow) -> str:
    bodies: list[str] = []
    features = join_items(example.keywords)
    for variant in range(3):
        intro = tone_intro(example, variant)
        extra = extra_sentence(example)
        short = sentence(f"{intro}. Nội dung tập trung vào {features}")
        if extra and variant != 1:
            short = f"{short} {extra}"
        benefits = benefit_lines(example, variant)
        highlights = highlight_lines(example, variant)
        bodies.append(
            "\n".join(
                [
                    f"Mô tả ngắn: {short}",
                    "Lợi ích chính:",
                    f"* {benefits[0]}",
                    f"* {benefits[1]}",
                    f"* {benefits[2]}",
                    "Đặc điểm nổi bật:",
                    f"* {highlights[0]}",
                    f"* {highlights[1]}",
                    f"* {highlights[2]}",
                    f"Lời kêu gọi hành động: {cta_for(example, variant)}.",
                ]
            )
        )
    return versioned(bodies)


def social_hook(example: ExampleRow, variant: int) -> str:
    product = example.product
    f1 = example.keywords[0]
    hooks = {
        "urgent": [
            f"Đang cần {product}? Xem ngay trước khi bỏ lỡ thông tin đáng cân nhắc.",
            f"Hôm nay là lúc kiểm tra kỹ {f1} cho nhu cầu của bạn.",
            f"Đừng để lựa chọn phù hợp trôi qua khi bạn đã biết mình cần gì.",
        ],
        "professional": [
            f"Ba điểm cần xem trước khi chọn {product}.",
            f"Thông tin chính giúp đánh giá {product} rõ ràng hơn.",
            f"Một cách tiếp cận gọn cho quyết định mua {product}.",
        ],
        "friendly": [
            f"Có thể {product} là món bạn đang cần cho ngày nhẹ hơn.",
            f"Mình xem nhanh {product} qua vài điểm dễ hiểu nhé.",
            f"Nếu đang phân vân, bắt đầu từ nhu cầu thật của bạn là ổn nhất.",
        ],
        "luxury": [
            f"Một lựa chọn tinh tế bắt đầu từ chi tiết vừa vặn.",
            f"{cap_first(product)} dành cho trải nghiệm được cân nhắc kỹ.",
            f"Khi chất lượng cảm nhận quan trọng hơn lời giới thiệu ồn ào.",
        ],
        "humorous": [
            f"Giỏ hàng đang nhìn bạn, còn {product} thì có vài điểm đáng nghe.",
            f"Không cần họp khẩn, nhưng {product} đáng được xem qua.",
            f"Một món hợp nhu cầu có thể làm bạn bớt thở dài khi mua sắm.",
        ],
        "emotional": [
            f"Có những món nhỏ làm một ngày quen thuộc dễ chịu hơn.",
            f"Khi chọn đúng, bạn thấy yên tâm từ những chi tiết rất gần.",
            f"{cap_first(product)} dành cho khoảnh khắc bạn muốn chăm chút hơn cho mình.",
        ],
    }
    return hooks.get(example.tone, hooks["friendly"])[variant % 3]


def social_caption(example: ExampleRow, variant: int) -> str:
    product = example.product
    audience = example.target
    features = join_items(example.keywords)
    extra = extra_sentence(example)
    captions = {
        "urgent": [
            f"{cap_first(product)} phù hợp với {audience}. Điểm cần xem ngay là {features}. {extra} Nếu các tiêu chí này đúng với nhu cầu của bạn, hãy kiểm tra sản phẩm trong hôm nay.",
            f"Không cần chờ thêm khi thông tin đã rõ. {product} có các điểm chính gồm {features}, giúp bạn quyết định nhanh hơn mà vẫn dựa trên thông tin đã có.",
            f"Nhu cầu đang có thì nên xử lý sớm. Với {audience}, {product} tập trung vào {features} và lời kêu gọi hành động rõ ràng.",
        ],
        "professional": [
            f"{cap_first(product)} được giới thiệu theo hướng rõ ràng cho {audience}. Nội dung nhấn vào {features}, tránh phóng đại và giúp người mua có cơ sở so sánh.",
            f"Nếu cần đánh giá nhanh, hãy bắt đầu từ ba điểm chính của {product}: {features}. {extra} Cách trình bày này phù hợp cho quyết định mua có kiểm soát.",
            f"Bài viết tập trung vào thông tin thực tế, nhóm người dùng phù hợp và bước hành động tiếp theo. {cap_first(product)} được đặt trong bối cảnh {audience}.",
        ],
        "friendly": [
            f"{cap_first(product)} hợp với {audience} đang muốn chọn một món dễ hiểu, dễ hỏi thêm. Bạn có thể xem trước {features}; nếu thấy đúng nhu cầu thì nhắn shop nhé.",
            f"Mình thích cách bắt đầu từ điều cần dùng thật. Với {product}, các điểm như {features} giúp bạn dễ hình dung hơn trước khi quyết định.",
            f"Không cần đọc quá nhiều lời quảng cáo. Chỉ cần xem {product} có hợp với {audience} không, rồi hỏi thêm nếu bạn muốn chọn kỹ hơn.",
        ],
        "luxury": [
            f"{cap_first(product)} được đặt trong một thông điệp tiết chế hơn, dành cho {audience} coi trọng trải nghiệm và sự chỉn chu. Các điểm chính gồm {features}.",
            f"Một lựa chọn cao cấp không cần nói quá nhiều. {product} thuyết phục bằng chi tiết đã có và cách trình bày tinh gọn.",
            f"Với {audience}, {product} hướng đến cảm giác chọn đúng, dùng vừa và có thông tin đủ rõ để quyết định.",
        ],
        "humorous": [
            f"{cap_first(product)} không hứa biến ngày của bạn thành phim quảng cáo, nhưng có {features}. Với {audience}, vậy là đủ để xem kỹ hơn rồi.",
            f"Nếu bạn đang lướt qua hàng loạt lựa chọn, {product} xin được nói ngắn gọn: có {features}, và có thể hợp với điều bạn cần.",
            f"Mua sắm vui hơn khi thông tin rõ. {product} dành cho {audience}, tập trung vào nhu cầu thật chứ không làm quá câu chuyện.",
        ],
        "emotional": [
            f"{cap_first(product)} dành cho {audience} muốn thấy yên tâm hơn trong lựa chọn nhỏ hằng ngày. Những điểm như {features} giúp quyết định trở nên nhẹ nhàng hơn.",
            f"Có lúc một sản phẩm phù hợp không cần nói lớn, chỉ cần đúng với điều bạn đang cần. {product} được gợi ý qua {features}.",
            f"Khi nhu cầu thật được lắng nghe, việc chọn mua cũng bớt căng. {product} đặt trọng tâm vào {audience} và những chi tiết đủ dùng.",
        ],
    }
    return captions.get(example.tone, captions["friendly"])[variant % 3]


def build_social_output(example: ExampleRow) -> str:
    bodies: list[str] = []
    for variant in range(3):
        bodies.append(
            "\n".join(
                [
                    f"Hook: {social_hook(example, variant)}",
                    f"Caption: {social_caption(example, variant)}",
                    f"Lời kêu gọi hành động: {cta_for(example, variant)}.",
                    f"Hashtags: {make_hashtags(example)}",
                ]
            )
        )
    return versioned(bodies)


def email_subject(example: ExampleRow, variant: int) -> str:
    product = example.product
    subjects = {
        "urgent": [
            f"Xem nhanh {product} cho nhu cầu hiện tại",
            f"Đừng bỏ qua các điểm chính của {product}",
            f"Hôm nay, bạn có thể kiểm tra {product} kỹ hơn",
        ],
        "professional": [
            f"Thông tin cần xem trước khi chọn {product}",
            f"Đánh giá {product} qua các tiêu chí chính",
            f"Gợi ý {product} cho nhu cầu sử dụng rõ ràng",
        ],
        "friendly": [
            f"Bạn đang xem {product}? Có vài điểm nên biết",
            f"Gợi ý nhẹ nhàng về {product}",
            f"Cùng xem {product} có hợp với bạn không",
        ],
        "luxury": [
            f"Một lựa chọn tinh tế hơn với {product}",
            f"Khám phá {product} qua những chi tiết chọn lọc",
            f"{cap_first(product)} cho trải nghiệm được cân nhắc kỹ",
        ],
        "humorous": [
            f"{cap_first(product)} có vài điểm muốn tự giới thiệu",
            f"Trước khi giỏ hàng thở dài, xem {product} nhé",
            f"Một email ngắn gọn về {product}",
        ],
        "emotional": [
            f"Một lựa chọn nhỏ để bạn yên tâm hơn",
            f"Khi {product} hợp với điều bạn đang cần",
            f"Gợi ý {product} cho ngày nhẹ nhàng hơn",
        ],
    }
    return subjects.get(example.tone, subjects["friendly"])[variant % 3]


def email_preview(example: ExampleRow, variant: int) -> str:
    features = join_items(example.keywords)
    previews = {
        "urgent": [
            f"Các điểm chính gồm {features}, nên xem sớm nếu đúng nhu cầu.",
            "Thông tin ngắn gọn để bạn quyết định nhanh hơn.",
            "Một lời nhắc kịp lúc cho lựa chọn đang cần.",
        ],
        "professional": [
            f"Tập trung vào {features} và nhóm người dùng phù hợp.",
            "Nội dung rõ ràng, gọn và dễ đối chiếu.",
            "Giúp bạn đánh giá trước khi đặt hàng.",
        ],
        "friendly": [
            "Mình tóm gọn vài điểm để bạn xem dễ hơn.",
            f"Nếu cần {features}, email này sẽ hữu ích.",
            "Đọc nhanh rồi hỏi thêm khi bạn cần nhé.",
        ],
        "luxury": [
            "Một cách nhìn tiết chế và chỉn chu hơn.",
            f"Tập trung vào {features} trong trải nghiệm sử dụng.",
            "Thông tin ngắn, chọn lọc và vừa đủ.",
        ],
        "humorous": [
            "Không dài dòng, chỉ vài điểm đáng xem.",
            f"Có {features}, và không có lời hứa quá tay.",
            "Một email nhẹ nhàng cho quyết định bớt rối.",
        ],
        "emotional": [
            "Vài thông tin giúp lựa chọn của bạn nhẹ lòng hơn.",
            f"Những điểm như {features} có thể làm bạn yên tâm hơn.",
            "Một gợi ý vừa phải cho nhu cầu hằng ngày.",
        ],
    }
    return previews.get(example.tone, previews["friendly"])[variant % 3]


def email_body(example: ExampleRow, variant: int) -> str:
    product = example.product
    audience = example.target
    features = join_items(example.keywords)
    extra = extra_sentence(example)
    body_by_tone = {
        "urgent": [
            f"Nếu bạn đang cần {product} cho {audience}, hãy xem ngay các điểm chính gồm {features}. {extra} Khi thông tin đã đủ rõ, bạn có thể chuyển sang bước đặt hàng hoặc hỏi shop để được hỗ trợ nhanh.",
            f"Nhu cầu phù hợp không nên bị để quá lâu. {cap_first(product)} tập trung vào {features}, giúp bạn kiểm tra nhanh và quyết định trong hôm nay nếu sản phẩm đúng với cách dùng của mình.",
            f"Email này tóm gọn những gì cần xem về {product}: {features}. Với {audience}, đây là nhóm thông tin nên được kiểm tra sớm trước khi lựa chọn khác làm bạn phân tán.",
        ],
        "professional": [
            f"{cap_first(product)} phù hợp để {audience} đánh giá theo các tiêu chí rõ ràng gồm {features}. {extra} Nội dung này giúp bạn xem lợi ích, đặc điểm và bước hành động tiếp theo trước khi đặt hàng.",
            f"Trước khi chọn {product}, bạn có thể đối chiếu ba thông tin chính: {features}. Cách tiếp cận này giúp quyết định mua hàng bớt cảm tính và dễ kiểm soát hơn.",
            f"Sản phẩm được giới thiệu cho {audience} với trọng tâm là thông tin thực tế. Không cần lời hứa quá mức, {product} nên được cân nhắc dựa trên {features}.",
        ],
        "friendly": [
            f"Nếu bạn đang xem {product}, mình gợi ý bắt đầu từ {features}. {extra} Những điểm này đủ để bạn hình dung sản phẩm có hợp với nhu cầu của mình không.",
            f"Mua sắm dễ hơn khi thông tin được nói thẳng và gần gũi. {cap_first(product)} dành cho {audience}, nhất là khi bạn muốn hỏi thêm trước khi quyết định.",
            f"Bạn có thể xem {product} như một lựa chọn đáng cân nhắc nếu {features} đúng với điều bạn cần. Cần chọn kỹ hơn thì cứ nhắn shop để được gợi ý.",
        ],
        "luxury": [
            f"{cap_first(product)} được giới thiệu với nhịp điệu tiết chế cho {audience}. Các điểm như {features} giúp sản phẩm tạo cảm giác chỉn chu mà không cần cách bán hàng ồn ào.",
            f"Nếu bạn ưu tiên trải nghiệm được cân nhắc kỹ, {product} là lựa chọn nên xem qua {features}. {extra} Thông tin được giữ gọn để bạn tập trung vào điều đáng giá.",
            f"Một quyết định mua hàng tinh tế bắt đầu từ chi tiết đủ rõ. Với {product}, nội dung tập trung vào {features} và cảm giác vừa vặn khi sử dụng.",
        ],
        "humorous": [
            f"{cap_first(product)} không cần tự giới thiệu quá dài. Với {audience}, sản phẩm có các điểm chính là {features}. Nếu nghe hợp lý, bạn có thể xem tiếp trước khi giỏ hàng kịp đổi ý.",
            f"Email này không làm quá câu chuyện. {cap_first(product)} chỉ cần nói rõ {features}, rồi để bạn quyết định xem có hợp nhu cầu không.",
            f"Nếu việc chọn mua đang hơi rối, hãy bắt đầu từ ba điểm dễ nhớ của {product}: {features}. Rõ hơn rồi thì hỏi shop một câu là xong bước tiếp theo.",
        ],
        "emotional": [
            f"Có những lựa chọn nhỏ giúp một ngày quen thuộc dễ chịu hơn. {cap_first(product)} dành cho {audience}, với các điểm như {features} để bạn thấy yên tâm hơn trước khi mua.",
            f"Nếu bạn đang muốn chọn {product} một cách nhẹ nhàng, hãy bắt đầu từ những chi tiết thật: {features}. {extra} Chúng giúp quyết định bớt phân vân mà không cần lời quảng cáo quá lớn.",
            f"Khi sản phẩm chạm đúng nhu cầu, cảm giác an tâm đến rất tự nhiên. Với {audience}, {product} gợi ý một lựa chọn vừa đủ qua {features}.",
        ],
    }
    return body_by_tone.get(example.tone, body_by_tone["friendly"])[variant % 3]


def build_email_output(example: ExampleRow) -> str:
    bodies: list[str] = []
    for variant in range(3):
        bodies.append(
            "\n".join(
                [
                    f"Subject: {email_subject(example, variant)}",
                    f"Preview text: {email_preview(example, variant)}",
                    "Lời chào: Chào bạn,",
                    f"Nội dung chính: {email_body(example, variant)}",
                    f"Lời kêu gọi hành động: {cta_for(example, variant)}.",
                ]
            )
        )
    return versioned(bodies)


def cta_microcopy(example: ExampleRow, variant: int) -> str:
    product = example.product
    features = join_items(example.keywords)
    audience = example.target
    extra = extra_sentence(example)
    tone = example.tone
    copy_by_tone = {
        "urgent": [
            f"Xem nhanh {features} và quyết định trong hôm nay nếu đúng nhu cầu.",
            f"Phù hợp với {audience}; thông tin đã đủ để bạn hỏi shop ngay.",
            f"{extra or sentence('Nhận tư vấn nhanh trước khi chọn sản phẩm')}",
        ],
        "professional": [
            f"Tập trung vào {features} để bạn đánh giá rõ ràng trước khi mua.",
            f"Nhận tư vấn dựa trên nhu cầu của {audience}.",
            f"Xem các điểm chính của {product} trước khi ra quyết định.",
        ],
        "friendly": [
            f"Xem {features} rồi hỏi shop nếu bạn muốn chọn kỹ hơn.",
            f"Mình sẽ gợi ý theo nhu cầu thật của {audience}.",
            f"Một bước nhẹ nhàng để xem {product} có hợp với bạn không.",
        ],
        "luxury": [
            f"Khám phá {product} qua những chi tiết đã có và được trình bày chọn lọc.",
            f"Tư vấn theo nhu cầu sử dụng, giữ trải nghiệm mua hàng chỉn chu.",
            f"Tập trung vào {features} để chọn lựa tinh tế hơn.",
        ],
        "humorous": [
            f"Thông tin gọn đủ dùng, không bắt bạn đọc một bài quảng cáo dài.",
            f"Hỏi nhanh để khỏi chọn nhầm rồi lại nhìn giỏ hàng thở dài.",
            f"Nếu {features} đúng nhu cầu, món này đáng được xem tiếp.",
        ],
        "emotional": [
            f"Chọn một món phù hợp để nhu cầu hằng ngày nhẹ nhàng hơn.",
            f"Nhắn shop nếu bạn muốn thấy yên tâm hơn trước khi mua.",
            f"Xem kỹ {features} để chọn {product} bằng cảm giác chắc lòng hơn.",
        ],
    }
    return copy_by_tone.get(tone, copy_by_tone["friendly"])[variant % 3]


def build_cta_output(example: ExampleRow) -> str:
    bodies: list[str] = []
    for variant in range(3):
        bodies.append(
            "\n".join(
                [
                    f"Lời kêu gọi hành động chính: {cta_for(example, variant)}",
                    f"Microcopy: {cta_microcopy(example, variant)}",
                ]
            )
        )
    return versioned(bodies)


def build_landing_output(example: ExampleRow) -> str:
    features = join_items(example.keywords)
    benefits = benefit_lines(example, 0)
    extra = extra_sentence(example)
    body = "\n".join(
        [
            f"Hero headline: {cap_first(example.product)} cho {example.target}.",
            f"Hero subheadline: Tập trung vào {features} để người mua hiểu nhanh giá trị trước khi hành động.",
            "Lợi ích chính:",
            f"* {benefits[0]}",
            f"* {benefits[1]}",
            f"* {benefits[2]}",
            f"Nội dung chính: {tone_intro(example, 0)}. Nội dung landing page giữ thông tin gọn, rõ lợi ích và không thêm lời hứa ngoài thông tin đầu vào.",
            f"Bằng chứng thuyết phục: Các điểm được dùng làm cơ sở là {features}. {extra}",
            f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
        ]
    )
    return versioned([body])


def build_seo_output(example: ExampleRow) -> str:
    features = join_items(example.keywords)
    body = "\n".join(
        [
            f"SEO title: {cap_first(example.product)} cho {example.target}",
            f"Meta description: Tìm hiểu {example.product} qua {features}, phù hợp với {example.target} và có lời kêu gọi hành động rõ ràng.",
            f"H1: {cap_first(example.product)} có phù hợp với {example.target}?",
            f"Mở bài: Bài viết này giúp bạn xem {example.product} theo các tiêu chí chính từ thông tin đầu vào, không nhồi từ khóa và không thêm lời hứa quá mức.",
            "Dàn ý nội dung:",
            f"* Nhu cầu của {example.target} khi cân nhắc {example.product}.",
            f"* Các điểm chính gồm {features}.",
            f"* Cách chọn và bước hành động tiếp theo.",
            f"Lời kêu gọi hành động: {cta_for(example, 0)}.",
        ]
    )
    return versioned([body])


def build_review_output(example: ExampleRow) -> str:
    f1, f2, f3 = example.keywords[:3]
    body = "\n".join(
        [
            "Tên người đánh giá: Minh Anh",
            f"Bối cảnh sử dụng: Tôi cân nhắc {example.product} vì thuộc nhóm {example.target}.",
            f"Nội dung review: Sau khi xem kỹ thông tin, tôi chú ý nhất đến {f1}. {cap_first(f2)} cũng là điểm giúp tôi yên tâm hơn, dù tôi vẫn cần hỏi thêm shop về cách chọn đúng theo nhu cầu cá nhân.",
            f"Điểm nổi bật được nhắc đến: {cap_first(f3)} và cách thông tin sản phẩm được trình bày khá dễ hiểu.",
            f"Kết luận / Gợi ý: Nếu bạn có nhu cầu giống tôi, nên xem {example.product} và hỏi thêm trước khi quyết định.",
        ]
    )
    return versioned([body])


def build_output(example: ExampleRow) -> str:
    builders = {
        "headline": build_headline_output,
        "description": build_description_output,
        "social": build_social_output,
        "email": build_email_output,
        "cta": build_cta_output,
        "landing": build_landing_output,
        "seo": build_seo_output,
        "review": build_review_output,
    }
    builder = builders.get(example.type)
    if not builder:
        return example.output
    return builder(example)


def row_dicts_from_examples(examples: list[ExampleRow], outputs: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for example, output in zip(examples, outputs):
        rows.append(
            {
                "excel_row": str(example.excel_row),
                "input": example.input,
                "output": output,
                "industry": example.industry,
                "tone": example.tone,
                "type": example.type,
                "product": example.product,
            }
        )
    return rows


def count_duplicate_outputs(outputs: list[str]) -> tuple[int, int]:
    counts = Counter(normalize_for_compare(output) for output in outputs)
    duplicate_rows = sum(count for count in counts.values() if count > 1)
    duplicate_groups = sum(1 for count in counts.values() if count > 1)
    return duplicate_rows, duplicate_groups


def validate_no_forbidden_input_phrases(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    for row in rows:
        lowered = row["input"].lower()
        for phrase in FORBIDDEN_INPUT_PHRASES:
            if phrase.lower() in lowered:
                errors.append(f"Dòng {row['excel_row']}: input còn chứa '{phrase}'")
    return errors


def validate_no_forbidden_output_phrases(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    forbidden = FORBIDDEN_OUTPUT_PHRASES + FORBIDDEN_CLAIMS
    for row in rows:
        lowered = row["output"].lower()
        for phrase in forbidden:
            if phrase.lower() in lowered:
                errors.append(f"Dòng {row['excel_row']}: output còn chứa '{phrase}'")
        if row["tone"] == "urgent":
            for phrase in URGENT_WEAK_PHRASES:
                if phrase.lower() in lowered:
                    errors.append(f"Dòng {row['excel_row']}: urgent tone còn CTA/câu yếu '{phrase}'")
    return errors


def split_versions(output: str) -> list[tuple[int, str]]:
    pattern = re.compile(r"^Phiên bản (\d+):\s*$", flags=re.MULTILINE)
    matches = list(pattern.finditer(output.strip()))
    versions: list[tuple[int, str]] = []
    for idx, match in enumerate(matches):
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(output)
        versions.append((int(match.group(1)), output[start:end].strip()))
    return versions


def line_label(line: str) -> str | None:
    if not line or line.startswith("* "):
        return None
    match = re.match(r"^([^:]{1,80}):", line)
    if not match:
        return None
    return match.group(1).strip()


def validate_fields_for_segment(row: dict[str, str], segment: str, version_no: int) -> list[str]:
    errors: list[str] = []
    content_type = row["type"]
    allowed = TYPE_FIELDS.get(content_type, [])
    lines = [line.rstrip() for line in segment.split("\n") if line.strip()]
    labels = [label for label in (line_label(line) for line in lines) if label]
    if labels != allowed:
        errors.append(
            f"Dòng {row['excel_row']} phiên bản {version_no}: field sai thứ tự/thiếu/thừa. Có {labels}, cần {allowed}"
        )
        return errors

    for label in labels:
        if label not in allowed:
            errors.append(f"Dòng {row['excel_row']} phiên bản {version_no}: field ngoài format '{label}'")

    bullet_count = sum(1 for line in lines if line.startswith("* "))
    if content_type == "description" and bullet_count != 6:
        errors.append(f"Dòng {row['excel_row']} phiên bản {version_no}: description cần đúng 6 bullet, hiện có {bullet_count}")
    if content_type in {"landing", "seo"} and bullet_count != 3:
        errors.append(f"Dòng {row['excel_row']} phiên bản {version_no}: {content_type} cần đúng 3 bullet, hiện có {bullet_count}")
    if content_type == "social":
        hashtag_lines = [line for line in lines if line.startswith("Hashtags:")]
        if len(hashtag_lines) != 1:
            errors.append(f"Dòng {row['excel_row']} phiên bản {version_no}: social cần đúng 1 dòng Hashtags")
        else:
            hashtag_count = len(re.findall(r"#[\wÀ-ỹ]+", hashtag_lines[0], flags=re.UNICODE))
            if hashtag_count < 3 or hashtag_count > 6:
                errors.append(
                    f"Dòng {row['excel_row']} phiên bản {version_no}: Hashtags cần 3-6 hashtag, hiện có {hashtag_count}"
                )
    if content_type == "cta" and bullet_count != 0:
        errors.append(f"Dòng {row['excel_row']} phiên bản {version_no}: cta không được có bullet hoặc bài dài")
    return errors


def validate_output_format_by_type(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    for row in rows:
        content_type = row["type"]
        expected_count = TYPE_VERSION_COUNTS.get(content_type)
        if expected_count is None:
            errors.append(f"Dòng {row['excel_row']}: type không hợp lệ '{content_type}'")
            continue

        versions = split_versions(row["output"])
        expected_numbers = list(range(1, expected_count + 1))
        actual_numbers = [number for number, _ in versions]
        if actual_numbers != expected_numbers:
            errors.append(
                f"Dòng {row['excel_row']}: số phiên bản sai. Có {actual_numbers}, cần {expected_numbers}"
            )
            continue
        normalized_segments = [normalize_for_compare(segment) for _, segment in versions]
        if len(normalized_segments) != len(set(normalized_segments)):
            errors.append(f"Dòng {row['excel_row']}: có phiên bản bị trùng nội dung trong cùng một output")
        for number, segment in versions:
            errors.extend(validate_fields_for_segment(row, segment, number))
    return errors


def validate_no_duplicate_outputs(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    seen: dict[str, str] = {}
    for row in rows:
        key = normalize_for_compare(row["output"])
        if key in seen:
            errors.append(f"Dòng {row['excel_row']}: output trùng chính xác với dòng {seen[key]}")
        else:
            seen[key] = row["excel_row"]
    return errors


def first_cta(output: str) -> str:
    for line in output.split("\n"):
        if line.startswith("Lời kêu gọi hành động:") or line.startswith("Lời kêu gọi hành động chính:"):
            return normalize_for_compare(line)
    return ""


def validate_tone_difference_for_same_product_type(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    groups: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        groups[(normalize_for_compare(row["product"]), normalize_for_compare(row["type"]))].append(row)

    for (_product, _content_type), group_rows in groups.items():
        for idx, left in enumerate(group_rows):
            for right in group_rows[idx + 1 :]:
                if left["tone"] == right["tone"]:
                    continue
                left_output = normalize_for_compare(left["output"])
                right_output = normalize_for_compare(right["output"])
                if left_output == right_output:
                    errors.append(
                        f"Dòng {left['excel_row']} và {right['excel_row']}: cùng product+type nhưng output giống nhau"
                    )
                    continue
                ratio = SequenceMatcher(None, left_output, right_output).ratio()
                if ratio >= 0.92:
                    errors.append(
                        f"Dòng {left['excel_row']} và {right['excel_row']}: output khác tone quá giống nhau ({ratio:.2f})"
                    )
                if first_cta(left["output"]) and first_cta(left["output"]) == first_cta(right["output"]):
                    errors.append(
                        f"Dòng {left['excel_row']} và {right['excel_row']}: CTA đầu tiên giống nhau giữa các tone"
                    )
    return errors


def validate_category_values(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    for row in rows:
        if row["industry"] not in VALID_INDUSTRIES:
            errors.append(f"Dòng {row['excel_row']}: industry không hợp lệ '{row['industry']}'")
        if row["tone"] not in VALID_TONES:
            errors.append(f"Dòng {row['excel_row']}: tone không hợp lệ '{row['tone']}'")
        if row["type"] not in VALID_TYPES:
            errors.append(f"Dòng {row['excel_row']}: type không hợp lệ '{row['type']}'")
    return errors


def run_all_validations(rows: list[dict[str, str]], header_map: dict[str, int]) -> list[str]:
    errors: list[str] = []
    errors.extend(validate_required_columns(header_map))
    errors.extend(validate_category_values(rows))
    errors.extend(validate_no_forbidden_input_phrases(rows))
    errors.extend(validate_no_forbidden_output_phrases(rows))
    errors.extend(validate_output_format_by_type(rows))
    errors.extend(validate_no_duplicate_outputs(rows))
    errors.extend(validate_tone_difference_for_same_product_type(rows))
    return errors


def count_rows_with_forbidden_phrases(rows: list[dict[str, str]]) -> int:
    forbidden_input = [phrase.lower() for phrase in FORBIDDEN_INPUT_PHRASES]
    forbidden_output = [phrase.lower() for phrase in FORBIDDEN_OUTPUT_PHRASES + FORBIDDEN_CLAIMS]
    count = 0
    for row in rows:
        input_lower = row["input"].lower()
        output_lower = row["output"].lower()
        if any(phrase in input_lower for phrase in forbidden_input) or any(
            phrase in output_lower for phrase in forbidden_output
        ):
            count += 1
    return count


def copy_column_widths(source_ws, target_ws, header_map: dict[str, int]) -> None:
    for target_col_idx, column_name in enumerate(REQUIRED_COLUMNS, start=1):
        source_idx = header_map[column_name] + 1
        source_letter = source_ws.cell(row=1, column=source_idx).column_letter
        target_letter = target_ws.cell(row=1, column=target_col_idx).column_letter
        source_dimension = source_ws.column_dimensions.get(source_letter)
        width = source_dimension.width if source_dimension and source_dimension.width else None
        if width:
            target_ws.column_dimensions[target_letter].width = min(max(width, 14), 90)
        elif column_name == "output":
            target_ws.column_dimensions[target_letter].width = 90
        elif column_name == "input":
            target_ws.column_dimensions[target_letter].width = 65
        else:
            target_ws.column_dimensions[target_letter].width = 18


def save_clean_workbook(source_ws, header_map: dict[str, int], rows: list[dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    target_ws = workbook.active
    target_ws.title = source_ws.title[:31]
    target_ws.append(REQUIRED_COLUMNS)
    for row in rows:
        target_ws.append([row[column] for column in REQUIRED_COLUMNS])

    copy_column_widths(source_ws, target_ws, header_map)
    target_ws.freeze_panes = "A2"
    for cell in target_ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in target_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def reread_clean_rows(path: Path) -> tuple[list[dict[str, str]], dict[str, int]]:
    workbook = load_workbook(path, data_only=False)
    worksheet, header_map = find_target_sheet(workbook)
    rows: list[dict[str, str]] = []
    for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {column: safe_value(values, header_map[column]) for column in REQUIRED_COLUMNS}
        row["excel_row"] = str(excel_row)
        rows.append(row)
    return rows, header_map


def build_report(
    source_path: Path,
    output_path: Path,
    report_path: Path,
    sheet_name: str,
    total_rows: int,
    input_cleaned_count: int,
    output_rewritten_count: int,
    duplicate_before_rows: int,
    duplicate_before_groups: int,
    duplicate_after_rows: int,
    duplicate_after_groups: int,
    forbidden_rows_after: int,
    validation_errors: list[str],
    warnings: list[str],
) -> str:
    validation_status = "PASS" if not validation_errors else "FAIL"
    lines = [
        "# Brand Voice Dataset Cleaning Report",
        "",
        f"- File nguồn: `{source_path}`",
        f"- File sạch: `{output_path}`",
        f"- Sheet xử lý: `{sheet_name}`",
        f"- Validation: **{validation_status}**",
        "",
        "## Thống kê",
        "",
        f"- Tổng số dòng đã xử lý: {total_rows}",
        f"- Số input đã được làm sạch: {input_cleaned_count}",
        f"- Số output đã được viết lại: {output_rewritten_count}",
        f"- Số output trùng trước khi làm sạch: {duplicate_before_rows} dòng thuộc {duplicate_before_groups} nhóm",
        f"- Số output trùng sau khi làm sạch: {duplicate_after_rows} dòng thuộc {duplicate_after_groups} nhóm",
        f"- Số dòng còn chứa cụm từ cấm sau khi làm sạch: {forbidden_rows_after}",
        "",
        "## Validation",
        "",
    ]
    if validation_errors:
        lines.append("Các lỗi còn lại:")
        for error in validation_errors:
            lines.append(f"- {error}")
    else:
        lines.append("- Tất cả validation đều pass.")

    lines.extend(["", "## Cảnh báo", ""])
    if warnings:
        for warning in warnings:
            lines.append(f"- {warning}")
    else:
        lines.append("- Không có cảnh báo.")

    report_text = "\n".join(lines) + "\n"
    report_path.write_text(report_text, encoding="utf-8")
    return validation_status


def main() -> int:
    try:
        source_path = resolve_source_path()
        output_path = source_path.with_name(CLEANED_FILE)
        report_path = source_path.with_name(REPORT_FILE)

        workbook = load_workbook(source_path, data_only=False)
        source_ws, header_map = find_target_sheet(workbook)
        column_errors = validate_required_columns(header_map)
        if column_errors:
            raise ValueError("; ".join(column_errors))

        examples = read_examples(source_ws, header_map)
        original_outputs = [example.output for example in examples]
        duplicate_before_rows, duplicate_before_groups = count_duplicate_outputs(original_outputs)

        cleaned_outputs = [build_output(example) for example in examples]
        clean_rows = row_dicts_from_examples(examples, cleaned_outputs)

        input_cleaned_count = 0
        for example in examples:
            original_input = safe_value(
                tuple(source_ws.cell(row=example.excel_row, column=idx + 1).value for idx in range(source_ws.max_column)),
                header_map["input"],
            )
            if normalize_ws_text(original_input).strip() != example.input.strip():
                input_cleaned_count += 1

        output_rewritten_count = sum(
            1
            for original, cleaned in zip(original_outputs, cleaned_outputs)
            if normalize_ws_text(original).strip() != cleaned.strip()
        )

        pre_save_errors = run_all_validations(clean_rows, {column: idx for idx, column in enumerate(REQUIRED_COLUMNS)})
        warnings: list[str] = []
        if pre_save_errors:
            warnings.append("Validation trước khi ghi file còn lỗi; vẫn ghi report để dễ kiểm tra.")

        save_clean_workbook(source_ws, header_map, clean_rows, output_path)
        reread_rows, reread_header_map = reread_clean_rows(output_path)
        validation_errors = run_all_validations(reread_rows, reread_header_map)

        duplicate_after_rows, duplicate_after_groups = count_duplicate_outputs([row["output"] for row in reread_rows])
        forbidden_rows_after = count_rows_with_forbidden_phrases(reread_rows)

        for row in reread_rows:
            if row["industry"] not in VALID_INDUSTRIES or row["tone"] not in VALID_TONES or row["type"] not in VALID_TYPES:
                warnings.append(f"Dòng {row['excel_row']}: có giá trị phân loại không hợp lệ và đã được giữ nguyên để tránh sửa sai.")

        validation_status = build_report(
            source_path=source_path.resolve(),
            output_path=output_path.resolve(),
            report_path=report_path.resolve(),
            sheet_name=source_ws.title,
            total_rows=len(examples),
            input_cleaned_count=input_cleaned_count,
            output_rewritten_count=output_rewritten_count,
            duplicate_before_rows=duplicate_before_rows,
            duplicate_before_groups=duplicate_before_groups,
            duplicate_after_rows=duplicate_after_rows,
            duplicate_after_groups=duplicate_after_groups,
            forbidden_rows_after=forbidden_rows_after,
            validation_errors=validation_errors,
            warnings=warnings,
        )

        print(f"Cleaned file: {output_path.resolve()}")
        print(f"Report file: {report_path.resolve()}")
        print(f"Validation: {validation_status}")
        return 0 if validation_status == "PASS" else 1
    except Exception as exc:  # noqa: BLE001 - terminal workflow should report the blocking error.
        print(f"Validation: FAIL")
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
